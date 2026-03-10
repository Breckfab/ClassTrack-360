# ============================================================
# INICIO PARTE 1 DE 2 — ClassTrack 360 v297
# ============================================================

import streamlit as st
from supabase import create_client
import datetime
import calendar
import streamlit.components.v1 as components
import io
import random
import string
import json
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
try:
    import plotly.graph_objects as go
    PLOTLY_OK = True
except ImportError:
    PLOTLY_OK = False

st.set_page_config(page_title="ClassTrack 360 v297", layout="wide")

SUPABASE_URL = "https://tzevdylabtradqmcqldx.supabase.co"
SUPABASE_KEY = "sb_publishable_SVgeWB2OpcuC3rd6L6b8sg_EcYfgUir"
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

ANIO_ACTUAL = datetime.date.today().year
SEDES_UNIVERSITARIO = ['daguerre']

def es_sistema_universitario(sede):
    return sede.lower() in SEDES_UNIVERSITARIO

def init_state():
    defaults = {
        'user': None, 'sede_admin': None,
        'editando_bitacora': None, 'editando_alumno': None, 'editando_curso': None,
        'confirmar_reset': None,
        'cal_mes': datetime.date.today().month, 'cal_anio': datetime.date.today().year,
        'es_suplente': False,
        'prev_curso_alumnos': None, 'prev_curso_notas_ver': None,
        'prev_curso_notas_carga': None, 'prev_curso_busq_hist': None,
        'busq_alumno_val': '', 'busq_nota_val': '', 'filtro_estado_val': 'Todos',
        'busq_carga_val': '', 'busq_contenido_hist_val': '', 'tipo_prof_val': 'Todos',
        'pantalla_login': 'login',
        'editando_tarea': None,
        'editando_tarea_legacy': None,
        'hist_curso': 'Todos',
        'hist_contenido': '',
        'hist_tipo_prof': 'Todos',
        'hist_anio': datetime.date.today().year,
        'hist_desde': datetime.date(datetime.date.today().year, 1, 1),
        'hist_hasta': datetime.date.today(),
        'asist_estados': {},
        'backup_json_bytes': None,
        'backup_excel_bytes': None,
        'backup_log_id': None,
        'backup_generado': False,
        'mostrar_salir_backup': False,
        # Flags de éxito para formularios (se muestran fuera del form)
        'ok_clase_guardada': False,
        'ok_curso_creado': None,
        'ok_curso_editado': False,
        'ok_alumno_registrado': None,
        'ok_alumno_editado': False,
        'ok_nota_agregada': None,
        'ok_bitacora_editada': False,
        'ok_calendario_guardado': False,
        'ok_registro_cuenta': None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

def footer():
    st.markdown('<div class="footer-cr">&#174; Sistema diseñado y realizado por Fabián Belledi &nbsp;·&nbsp; 2026</div>', unsafe_allow_html=True)

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;700;800&family=DM+Mono:wght@300;400&display=swap');
    .stApp { background: #080b10; color: #e8eaf0; font-family: 'DM Mono', monospace; }
    .stApp::before {
        content: ''; position: fixed; inset: 0; z-index: 0; pointer-events: none;
        background-image: linear-gradient(rgba(79,172,254,0.04) 1px, transparent 1px), linear-gradient(90deg, rgba(79,172,254,0.04) 1px, transparent 1px);
        background-size: 48px 48px;
    }
    .planilla-row { background: rgba(255,255,255,0.03); padding: 15px; border-radius: 10px; margin-bottom: 10px; border-left: 5px solid #4facfe; border: 1px solid rgba(255,255,255,0.1); }
    .tarea-alerta { background: rgba(255,193,7,0.25); border: 2px solid #ffc107; padding: 20px; border-radius: 12px; color: #ffc107; text-align: center; font-weight: 800; margin-bottom: 10px; font-size: 1.2rem; box-shadow: 0 4px 15px rgba(0,0,0,0.5); }
    .tarea-card { background: rgba(255,193,7,0.08); border: 1px solid rgba(255,193,7,0.25); border-radius: 10px; padding: 14px 18px; margin-bottom: 8px; }
    .tarea-card .tarea-titulo { color: #ffc107; font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 4px; }
    .tarea-card .tarea-texto { color: #e8eaf0; font-size: 0.9rem; margin-bottom: 4px; }
    .tarea-card .tarea-fecha { color: #778; font-size: 0.75rem; }
    .tarea-card-done { background: rgba(79,172,254,0.05); border: 1px solid rgba(79,172,254,0.15); border-radius: 10px; padding: 10px 18px; margin-bottom: 8px; opacity: 0.5; }
    .tarea-card-done .tarea-titulo { color: #4facfe; font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 4px; }
    .tarea-card-done .tarea-texto { color: #778; font-size: 0.9rem; text-decoration: line-through; margin-bottom: 4px; }
    .tarea-card-done .tarea-fecha { color: #556; font-size: 0.75rem; }
    .tarea-card-vencida { background: rgba(255,77,109,0.10); border: 2px solid rgba(255,77,109,0.5); border-radius: 10px; padding: 14px 18px; margin-bottom: 8px; }
    .tarea-card-vencida .tarea-titulo { color: #ff4d6d; font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 4px; }
    .tarea-card-vencida .tarea-texto { color: #e8eaf0; font-size: 0.9rem; margin-bottom: 4px; }
    .tarea-card-vencida .tarea-fecha { color: #ff4d6d; font-size: 0.75rem; font-weight: 700; }
    .vencidas-header { color: #ff4d6d; font-weight: 700; font-size: 0.9rem; margin-bottom: 10px; margin-top: 4px; background: rgba(255,77,109,0.07); border: 1px solid rgba(255,77,109,0.25); border-radius: 8px; padding: 8px 14px; }
    .badge-vencidas { background: #ff4d6d; color: #fff; font-family: 'Syne', sans-serif; font-weight: 800; font-size: 0.8rem; border-radius: 20px; padding: 4px 14px; display: inline-block; margin-bottom: 8px; letter-spacing: 0.05em; }
    .badge-clases { background: rgba(79,172,254,0.15); border: 1px solid rgba(79,172,254,0.4); color: #4facfe; font-family: 'Syne', sans-serif; font-weight: 800; font-size: 0.8rem; border-radius: 20px; padding: 4px 14px; display: inline-block; margin-bottom: 8px; letter-spacing: 0.05em; }
    .proxima-clase-card { background: rgba(79,172,254,0.07); border: 1px solid rgba(79,172,254,0.25); border-radius: 10px; padding: 12px 18px; margin-bottom: 12px; }
    .proxima-clase-card .pc-titulo { color: #4facfe; font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 4px; }
    .proxima-clase-card .pc-fecha { color: #e8eaf0; font-size: 1rem; font-weight: 700; margin-bottom: 2px; }
    .proxima-clase-card .pc-detalle { color: #778; font-size: 0.78rem; }
    .contador-card { background: rgba(79,172,254,0.05); border: 1px solid rgba(79,172,254,0.2); border-radius: 12px; padding: 18px 22px; margin-bottom: 12px; }
    .contador-card .cc-nombre { color: #e8eaf0; font-size: 0.95rem; font-weight: 700; margin-bottom: 8px; }
    .contador-card .cc-info { color: #4facfe; font-size: 0.78rem; margin-bottom: 4px; }
    .contador-card .cc-clases { color: #a0c4ff; font-size: 0.82rem; margin-top: 6px; }
    .asist-card { background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.08); border-radius: 12px; padding: 14px 18px; margin-bottom: 8px; display: flex; align-items: center; justify-content: space-between; }
    .asist-nombre { color: #e8eaf0; font-size: 0.9rem; font-weight: 600; }
    .asist-presente { color: #4facfe; background: rgba(79,172,254,0.1); border: 1px solid rgba(79,172,254,0.3); padding: 3px 12px; border-radius: 20px; font-size: 0.78rem; font-weight: 700; }
    .asist-tarde { color: #ffc107; background: rgba(255,193,7,0.1); border: 1px solid rgba(255,193,7,0.3); padding: 3px 12px; border-radius: 20px; font-size: 0.78rem; font-weight: 700; }
    .asist-ausente { color: #ff4d6d; background: rgba(255,77,109,0.1); border: 1px solid rgba(255,77,109,0.3); padding: 3px 12px; border-radius: 20px; font-size: 0.78rem; font-weight: 700; }
    .resumen-asist { background: rgba(79,172,254,0.05); border: 1px solid rgba(79,172,254,0.15); border-radius: 10px; padding: 12px 18px; margin-bottom: 10px; }
    .resumen-asist-titulo { color: #4facfe; font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 8px; }
    .resumen-fila { display: flex; justify-content: space-between; padding: 3px 0; border-bottom: 1px solid rgba(255,255,255,0.04); font-size: 0.82rem; color: #99a; }
    .resumen-fila:last-child { border-bottom: none; }
    .stat-card { background: rgba(79,172,254,0.1); border: 1px solid #4facfe; padding: 10px; border-radius: 8px; text-align: center; margin-bottom: 10px; }
    .nota-existente { color: #4facfe; font-size: 0.85rem; margin-top: 4px; }
    .alumno-block { background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.08); border-radius: 12px; padding: 14px 18px; margin-bottom: 12px; }
    .alumno-block .nombre { font-weight: 600; color: #e8eaf0; font-size: 0.95rem; margin-bottom: 10px; }
    .nota-linea { display: flex; justify-content: space-between; align-items: center; padding: 4px 0; border-bottom: 1px solid rgba(255,255,255,0.05); font-size: 0.85rem; color: #99a; }
    .nota-linea:last-of-type { border-bottom: none; }
    .nota-badge { font-family: 'Syne', sans-serif; font-weight: 700; font-size: 0.85rem; padding: 2px 8px; border-radius: 5px; }
    .nota-badge.baja { color: #ff4d6d; background: rgba(255,77,109,0.1); border: 1px solid rgba(255,77,109,0.3); }
    .nota-badge.media { color: #ffc107; background: rgba(255,193,7,0.1); border: 1px solid rgba(255,193,7,0.3); }
    .nota-badge.alta { color: #4facfe; background: rgba(79,172,254,0.1); border: 1px solid rgba(79,172,254,0.3); }
    .promedio-linea { display: flex; justify-content: space-between; align-items: center; margin-top: 8px; padding-top: 8px; border-top: 1px solid rgba(79,172,254,0.2); font-size: 0.85rem; color: #4facfe; font-weight: 600; }
    .aprobado { color: #4facfe; background: rgba(79,172,254,0.1); border: 1px solid rgba(79,172,254,0.3); padding: 2px 10px; border-radius: 5px; font-size: 0.8rem; font-weight: 700; }
    .desaprobado { color: #ff4d6d; background: rgba(255,77,109,0.1); border: 1px solid rgba(255,77,109,0.3); padding: 2px 10px; border-radius: 5px; font-size: 0.8rem; font-weight: 700; }
    .biblio-box { background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.06); border-radius: 8px; padding: 10px 14px; margin-top: 6px; font-size: 0.8rem; color: #778; }
    .suplente-badge { background: rgba(255,193,7,0.12); border: 1px solid rgba(255,193,7,0.35); border-radius: 6px; padding: 3px 10px; color: #ffc107; font-size: 0.78rem; font-weight: 700; display: inline-block; margin-top: 4px; }
    .titular-badge { background: rgba(79,172,254,0.1); border: 1px solid rgba(79,172,254,0.3); border-radius: 6px; padding: 3px 10px; color: #4facfe; font-size: 0.78rem; font-weight: 700; display: inline-block; margin-top: 4px; }
    .filtros-box { background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.07); border-radius: 10px; padding: 14px 18px; margin-bottom: 16px; }
    .no-encontrado { background: rgba(255,77,109,0.07); border: 1px solid rgba(255,77,109,0.2); border-radius: 8px; padding: 12px 18px; color: #ff4d6d; font-size: 0.88rem; margin-top: 8px; }
    .cal-box { background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.07); border-radius: 10px; padding: 10px; margin-top: 8px; }
    .cal-titulo { text-align: center; font-family: 'Syne', sans-serif; font-weight: 700; font-size: 0.85rem; color: #4facfe; margin-bottom: 8px; }
    .cal-tabla { width: 100%; border-collapse: collapse; font-size: 0.72rem; }
    .cal-tabla th { color: #4facfe; text-align: center; padding: 2px; font-weight: 600; }
    .cal-tabla td { text-align: center; padding: 3px 2px; color: #99a; }
    .cal-tabla td.hoy { background: #4facfe; color: #080b10; border-radius: 50%; font-weight: 800; }
    .cal-tabla td.vacio { color: transparent; }
    .admin-badge { background: rgba(255,77,109,0.15); border: 1px solid rgba(255,77,109,0.4); border-radius: 8px; padding: 8px 14px; color: #ff4d6d; font-size: 0.8rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; display: inline-block; margin-bottom: 16px; }
    .reset-box { background: rgba(255,77,109,0.05); border: 2px solid rgba(255,77,109,0.3); border-radius: 12px; padding: 20px; margin-top: 20px; }
    .reset-titulo { color: #ff4d6d; font-family: 'Syne', sans-serif; font-weight: 700; font-size: 1rem; margin-bottom: 8px; }
    .advertencia-box { background: rgba(255,193,7,0.1); border: 2px solid #ffc107; border-radius: 10px; padding: 16px; margin: 16px 0; color: #ffc107; font-size: 0.9rem; }
    .login-box { background: rgba(255,255,255,0.03); border: 1px solid rgba(79,172,254,0.2); border-radius: 16px; padding: 40px; margin-top: 60px; }
    .login-logo { font-family: 'Syne', sans-serif; font-weight: 800; font-size: 1.4rem; letter-spacing: 0.05em; color: #e8eaf0; margin-bottom: 6px; }
    .login-logo span { color: #4facfe; }
    .login-eyebrow { font-size: 0.72rem; text-transform: uppercase; letter-spacing: 0.15em; color: #4facfe; margin-bottom: 16px; }
    .login-title { font-family: 'Syne', sans-serif; font-weight: 700; font-size: 1.6rem; margin-bottom: 28px; color: #e8eaf0; }
    .login-footer { font-size: 0.72rem; color: #3a4358; text-align: center; margin-top: 24px; border-top: 1px solid rgba(255,255,255,0.05); padding-top: 16px; }
    .codigo-box { background: rgba(79,172,254,0.08); border: 1px solid rgba(79,172,254,0.3); border-radius: 8px; padding: 12px 18px; font-family: 'DM Mono', monospace; font-size: 1.1rem; color: #4facfe; font-weight: 700; letter-spacing: 0.15em; text-align: center; margin: 8px 0; }
    .habilitado-tag { color: #4facfe; background: rgba(79,172,254,0.1); border: 1px solid rgba(79,172,254,0.3); padding: 2px 10px; border-radius: 5px; font-size: 0.75rem; font-weight: 700; }
    .deshabilitado-tag { color: #ff4d6d; background: rgba(255,77,109,0.1); border: 1px solid rgba(255,77,109,0.3); padding: 2px 10px; border-radius: 5px; font-size: 0.75rem; font-weight: 700; }
    .footer-cr { position: fixed; bottom: 0; left: 0; right: 0; text-align: center; padding: 6px; font-size: 0.72rem; color: #c8d8f0; font-family: 'DM Mono', monospace; background: #080b10; border-top: 1px solid rgba(255,255,255,0.08); z-index: 999; letter-spacing: 0.05em; }
    .tareas-pendientes-header { color: #ffc107; font-weight: 700; font-size: 0.9rem; margin-bottom: 10px; margin-top: 4px; }
    .registro-link { text-align: center; margin-top: 16px; font-size: 0.78rem; color: #4a5568; }
    .email-tag { color: #4facfe; font-size: 0.78rem; margin-top: 2px; }
    .calendario-box { background: rgba(79,172,254,0.06); border: 1px solid rgba(79,172,254,0.2); border-radius: 12px; padding: 16px 20px; margin-bottom: 16px; }
    .calendario-box .cal-header { color: #4facfe; font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 10px; }
    .calendario-box .cal-fechas { color: #e8eaf0; font-size: 0.9rem; }
    .calendario-box .cal-dias { color: #778; font-size: 0.78rem; margin-top: 4px; }
    .cronograma-box { background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.07); border-radius: 10px; padding: 14px 18px; margin-bottom: 12px; }
    .cronograma-titulo { color: #4facfe; font-size: 0.8rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 8px; }
    .backup-box { background: rgba(79,172,254,0.05); border: 1px solid rgba(79,172,254,0.2); border-radius: 14px; padding: 20px 24px; margin-bottom: 16px; }
    .backup-titulo { color: #4facfe; font-family: 'Syne', sans-serif; font-weight: 700; font-size: 1rem; margin-bottom: 8px; }
    .backup-aviso { background: rgba(255,193,7,0.08); border: 1px solid rgba(255,193,7,0.3); border-radius: 10px; padding: 14px 18px; margin-bottom: 16px; color: #ffc107; font-size: 0.88rem; }
    .backup-hist-row { display: flex; justify-content: space-between; align-items: center; padding: 10px 0; border-bottom: 1px solid rgba(255,255,255,0.05); font-size: 0.82rem; color: #99a; }
    .backup-hist-row:last-child { border-bottom: none; }
    .backup-ok { color: #4facfe; font-weight: 700; }
    .backup-no { color: #556; }
    .restore-box { background: rgba(255,77,109,0.04); border: 2px solid rgba(255,77,109,0.3); border-radius: 14px; padding: 20px 24px; margin-top: 24px; }
    .restore-titulo { color: #ff4d6d; font-family: 'Syne', sans-serif; font-weight: 700; font-size: 1rem; margin-bottom: 8px; }
    .salir-backup-overlay { background: rgba(8,11,16,0.97); border: 2px solid rgba(79,172,254,0.4); border-radius: 16px; padding: 28px 32px; margin: 20px 0; text-align: center; }
    .salir-backup-titulo { color: #4facfe; font-family: 'Syne', sans-serif; font-weight: 800; font-size: 1.1rem; margin-bottom: 12px; }
    .salir-backup-info { color: #e8eaf0; font-size: 0.9rem; margin-bottom: 20px; }
    </style>
""", unsafe_allow_html=True)

# =========================================================
# --- FUNCIONES GENERALES ---
# =========================================================
def no_encontrado(msg):
    st.markdown(f'<div class="no-encontrado">🔍 {msg}</div>', unsafe_allow_html=True)

def color_nota(n):
    if n <= 3: return "baja"
    elif n <= 5: return "media"
    else: return "alta"

def estado_aprobacion(promedio, nota_aprobacion):
    if promedio is None or nota_aprobacion is None: return ""
    if promedio >= float(nota_aprobacion): return '<span class="aprobado">✅ APROBADO</span>'
    else: return '<span class="desaprobado">❌ DESAPROBADO</span>'

def estado_texto(promedio, nota_aprobacion):
    if promedio is None or nota_aprobacion is None: return "-"
    return "APROBADO" if promedio >= float(nota_aprobacion) else "DESAPROBADO"

def validar_hora(h):
    try:
        datetime.datetime.strptime(h.strip(), "%H:%M")
        return True
    except: return False

def format_horario(inicio, fin):
    if inicio and fin: return f"{str(inicio)[:5]} → {str(fin)[:5]}"
    elif inicio: return str(inicio)[:5]
    return "-"

def generar_codigo():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))

def marcar_tarea(bit_id, num_tarea, completada):
    try:
        supabase.table("bitacora").update({f"tarea{num_tarea}_completada": completada}).eq("id", bit_id).execute()
        st.rerun()
    except Exception as e:
        st.error(f"Error: {e}")

def marcar_tarea_proxima(bit_id, completada):
    try:
        supabase.table("bitacora").update({"tarea_proxima_completada": completada}).eq("id", bit_id).execute()
        st.rerun()
    except Exception as e:
        st.error(f"Error: {e}")

def guardar_edicion_tarea(bit_id, num_tarea, nuevo_texto, nueva_fecha):
    try:
        supabase.table("bitacora").update({
            f"tarea{num_tarea}": nuevo_texto.strip() or None,
            f"tarea{num_tarea}_fecha": str(nueva_fecha) if nuevo_texto.strip() else None,
        }).eq("id", bit_id).execute()
        st.session_state.editando_tarea = None
        st.rerun()
    except Exception as e:
        st.error(f"Error: {e}")

def guardar_edicion_tarea_legacy(bit_id, nuevo_texto):
    try:
        supabase.table("bitacora").update({
            "tarea_proxima": nuevo_texto.strip() or None,
        }).eq("id", bit_id).execute()
        st.session_state.editando_tarea_legacy = None
        st.rerun()
    except Exception as e:
        st.error(f"Error: {e}")

def get_tareas_vencidas_count(profesor_id):
    hoy = datetime.date.today()
    try:
        res_insc = supabase.table("inscripciones").select("id").eq("profesor_id", profesor_id).is_("alumno_id", "null").execute()
        if not res_insc.data: return 0
        ids = [r['id'] for r in res_insc.data]
        total_vencidas = 0
        for iid in ids:
            res = supabase.table("bitacora").select(
                "tarea1, tarea1_fecha, tarea1_completada, tarea2, tarea2_fecha, tarea2_completada, tarea3, tarea3_fecha, tarea3_completada"
            ).eq("inscripcion_id", iid).execute()
            for reg in (res.data or []):
                for i in range(1, 4):
                    txt = reg.get(f'tarea{i}')
                    fecha_t = reg.get(f'tarea{i}_fecha')
                    completada = reg.get(f'tarea{i}_completada', False)
                    if txt and not completada and fecha_t:
                        if datetime.date.fromisoformat(fecha_t) < hoy:
                            total_vencidas += 1
        return total_vencidas
    except: return 0

def get_stats_curso(inscripcion_id):
    try:
        res = supabase.table("bitacora").select("fecha").eq("inscripcion_id", inscripcion_id).order("fecha", desc=True).execute()
        cant = len(res.data) if res.data else 0
        ultima = datetime.date.fromisoformat(res.data[0]['fecha']).strftime('%d/%m/%Y') if res.data else None
        return cant, ultima
    except:
        return 0, None

def get_total_clases_sidebar(profesor_id):
    try:
        res_insc = supabase.table("inscripciones").select("id").eq("profesor_id", profesor_id).is_("alumno_id", "null").execute()
        if not res_insc.data: return 0
        total = 0
        for r in res_insc.data:
            cant, _ = get_stats_curso(r['id'])
            total += cant
        return total
    except: return 0

DIAS_SEMANA_MAP = {
    'lunes': 0, 'martes': 1, 'miércoles': 2, 'miercoles': 2,
    'jueves': 3, 'viernes': 4, 'sábado': 5, 'sabado': 5, 'domingo': 6
}
DIAS_SEMANA_ES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
MESES_ES_LARGO = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto','septiembre','octubre','noviembre','diciembre']
MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

def extraer_dias_curso(nombre_curso):
    dias = []
    try:
        if '(' in nombre_curso and ')' in nombre_curso:
            parte = nombre_curso.split('(')[1].split(')')[0]
            for token in parte.split(','):
                t = token.strip().lower()
                if t in DIAS_SEMANA_MAP:
                    dias.append(DIAS_SEMANA_MAP[t])
    except: pass
    return sorted(set(dias))

def get_proxima_clase(nombre_curso, hora_inicio, hora_fin):
    dias = extraer_dias_curso(nombre_curso)
    if not dias: return None
    hoy = datetime.date.today()
    for offset in range(1, 8):
        candidato = hoy + datetime.timedelta(days=offset)
        if candidato.weekday() in dias:
            nombre_dia = DIAS_SEMANA_ES[candidato.weekday()]
            fecha_fmt = f"{nombre_dia} {candidato.day} de {MESES_ES_LARGO[candidato.month - 1]}"
            horario = format_horario(str(hora_inicio or '')[:5], str(hora_fin or '')[:5])
            return {'fecha_fmt': fecha_fmt, 'dias_faltan': offset, 'horario': horario, 'fecha': candidato}
    return None

def render_calendario(mes, anio):
    hoy = datetime.date.today()
    cal = calendar.monthcalendar(anio, mes)
    filas = ""
    for semana in cal:
        fila = ""
        for dia in semana:
            if dia == 0: fila += '<td class="vacio">·</td>'
            elif dia == hoy.day and mes == hoy.month and anio == hoy.year: fila += f'<td class="hoy">{dia}</td>'
            else: fila += f'<td>{dia}</td>'
        filas += f"<tr>{fila}</tr>"
    return f"""<div class="cal-box"><div class="cal-titulo">{MESES_ES[mes-1]} {anio}</div>
    <table class="cal-tabla"><tr><th>Lu</th><th>Ma</th><th>Mi</th><th>Ju</th><th>Vi</th><th>Sa</th><th>Do</th></tr>{filas}</table></div>"""

def get_calendario_sede(sede):
    try:
        res = supabase.table("calendario_sede").select("*").eq("sede", sede).execute()
        return res.data[0] if res.data else None
    except: return None

def upsert_calendario_sede(sede, fecha_inicio, fecha_fin):
    try:
        existing = get_calendario_sede(sede)
        if existing:
            supabase.table("calendario_sede").update({
                "fecha_inicio": str(fecha_inicio) if fecha_inicio else None,
                "fecha_fin": str(fecha_fin) if fecha_fin else None,
                "updated_at": datetime.datetime.now().isoformat()
            }).eq("sede", sede).execute()
        else:
            supabase.table("calendario_sede").insert({
                "sede": sede,
                "fecha_inicio": str(fecha_inicio) if fecha_inicio else None,
                "fecha_fin": str(fecha_fin) if fecha_fin else None,
            }).execute()
        return True
    except Exception as e:
        st.error(f"Error guardando calendario: {e}")
        return False

def subir_cronograma(sede, cuatrimestre, archivo):
    try:
        ext = archivo.name.split('.')[-1].lower()
        path = f"{sede}/cronograma_{cuatrimestre}c.{ext}"
        data = archivo.read()
        try: supabase.storage.from_("cronogramas").remove([path])
        except: pass
        supabase.storage.from_("cronogramas").upload(path, data, {"content-type": archivo.type, "upsert": "true"})
        url = supabase.storage.from_("cronogramas").get_public_url(path)
        existing = get_calendario_sede(sede)
        campo_url = f"cronograma_{cuatrimestre}c_url"
        campo_nombre = f"cronograma_{cuatrimestre}c_nombre"
        if existing:
            supabase.table("calendario_sede").update({
                campo_url: url, campo_nombre: archivo.name,
                "updated_at": datetime.datetime.now().isoformat()
            }).eq("sede", sede).execute()
        else:
            supabase.table("calendario_sede").insert({
                "sede": sede, campo_url: url, campo_nombre: archivo.name,
            }).execute()
        return url
    except Exception as e:
        st.error(f"Error subiendo cronograma: {e}")
        return None

def render_cronograma_visor(url, nombre, titulo):
    if not url: return
    ext = nombre.split('.')[-1].lower() if nombre else ''
    st.markdown(f'<div class="cronograma-box"><div class="cronograma-titulo">📅 {titulo}</div>', unsafe_allow_html=True)
    if ext in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
        st.image(url, use_container_width=True)
    elif ext == 'pdf':
        components.html(f'<iframe src="{url}" width="100%" height="600px" style="border:none;border-radius:8px;"></iframe>', height=620)
    else:
        st.markdown(f'⚠️ Formato {ext.upper()} — <a href="{url}" target="_blank" style="color:#4facfe;">Descargar archivo</a>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

def reset_completo_sede(sede_nombre):
    try:
        res_u = supabase.table("usuarios").select("id").eq("sede", sede_nombre).execute()
        if not res_u.data: return False, "No se encontró la sede."
        profesor_id = res_u.data[0]['id']
        res_insc = supabase.table("inscripciones").select("id, alumno_id").eq("profesor_id", profesor_id).execute()
        ids_alumnos = list(set([i['alumno_id'] for i in res_insc.data if i.get('alumno_id')] if res_insc.data else []))
        for iid in [i['id'] for i in res_insc.data] if res_insc.data else []:
            supabase.table("notas").delete().eq("inscripcion_id", iid).execute()
            supabase.table("bitacora").delete().eq("inscripcion_id", iid).execute()
            supabase.table("asistencia").delete().eq("inscripcion_id", iid).execute()
        supabase.table("inscripciones").delete().eq("profesor_id", profesor_id).execute()
        for aid in ids_alumnos:
            supabase.table("alumnos").delete().eq("id", aid).execute()
        return True, f"Sede {sede_nombre.upper()} reseteada correctamente."
    except Exception as e:
        return False, str(e)

# =========================================================
# --- FUNCIONES DE ASISTENCIA ---
# =========================================================
def get_cuatrimestre_actual():
    mes = datetime.date.today().month
    return 1 if mes <= 6 else 2

def guardar_asistencia_instituto(inscripcion_id, fecha, estados_dict):
    try:
        supabase.table("asistencia").delete().eq("inscripcion_id", inscripcion_id).eq("fecha", str(fecha)).execute()
        rows = []
        for alumno_id, estado in estados_dict.items():
            rows.append({
                "inscripcion_id": inscripcion_id,
                "alumno_id": alumno_id,
                "fecha": str(fecha),
                "estado": estado,
                "materia": "",
                "observacion": None,
            })
        if rows:
            supabase.table("asistencia").insert(rows).execute()
        return True
    except Exception as e:
        st.error(f"Error guardando asistencia: {e}")
        return False

def guardar_asistencia_universitario(inscripcion_id, fecha, cuatrimestre, horas_catedra, estados_por_hora):
    try:
        supabase.table("asistencia").delete().eq("inscripcion_id", inscripcion_id).eq("fecha", str(fecha)).execute()
        rows = []
        for hora in range(1, horas_catedra + 1):
            estados = estados_por_hora.get(hora, {})
            for alumno_id, estado in estados.items():
                rows.append({
                    "inscripcion_id": inscripcion_id,
                    "alumno_id": alumno_id,
                    "fecha": str(fecha),
                    "estado": estado,
                    "cuatrimestre": cuatrimestre,
                    "hora_catedra": hora,
                    "materia": "",
                    "observacion": None,
                })
        if rows:
            supabase.table("asistencia").insert(rows).execute()
        return True
    except Exception as e:
        st.error(f"Error guardando asistencia: {e}")
        return False

def get_asistencia_fecha(inscripcion_id, fecha):
    try:
        res = supabase.table("asistencia").select("*").eq("inscripcion_id", inscripcion_id).eq("fecha", str(fecha)).execute()
        return res.data or []
    except: return []

def get_resumen_asistencia_instituto(inscripcion_id, mes, anio):
    try:
        desde = datetime.date(anio, mes, 1)
        ultimo_dia = calendar.monthrange(anio, mes)[1]
        hasta = datetime.date(anio, mes, ultimo_dia)
        res = supabase.table("asistencia").select("*").eq("inscripcion_id", inscripcion_id).gte("fecha", str(desde)).lte("fecha", str(hasta)).execute()
        resultado = {}
        for r in (res.data or []):
            aid = r['alumno_id']
            if aid not in resultado: resultado[aid] = {}
            resultado[aid][r['fecha']] = r['estado']
        return resultado
    except: return {}

def get_asistencia_anual_alumno(alumno_id, profesor_id, anio):
    """Trae toda la asistencia de un alumno en el año lectivo, en todos sus cursos."""
    try:
        desde = datetime.date(anio, 1, 1)
        hasta = datetime.date(anio, 12, 31)
        # Buscar todas las inscripciones del alumno con este profesor
        res_insc = supabase.table("inscripciones").select("id, nombre_curso_materia").eq("profesor_id", profesor_id).eq("alumno_id", alumno_id).execute()
        if not res_insc.data:
            return []
        resultado = []
        for insc in res_insc.data:
            res = supabase.table("asistencia").select("fecha, estado, hora_catedra").eq("inscripcion_id", insc['id']).eq("alumno_id", alumno_id).gte("fecha", str(desde)).lte("fecha", str(hasta)).order("fecha").execute()
            for r in (res.data or []):
                resultado.append({
                    'fecha': r['fecha'],
                    'estado': r['estado'],
                    'curso': insc['nombre_curso_materia'],
                    'hora_catedra': r.get('hora_catedra'),
                })
        resultado.sort(key=lambda x: x['fecha'])
        return resultado
    except: return []

def get_resumen_asistencia_universitario(inscripcion_id, cuatrimestre):
    try:
        res = supabase.table("asistencia").select("*").eq("inscripcion_id", inscripcion_id).eq("cuatrimestre", cuatrimestre).execute()
        resultado = {}
        for r in (res.data or []):
            aid = r['alumno_id']
            if aid not in resultado: resultado[aid] = {'presente': 0, 'tarde': 0, 'ausente': 0}
            estado = r.get('estado', 'presente')
            if estado in resultado[aid]: resultado[aid][estado] += 1
        return resultado
    except: return {}

def get_alumnos_curso(nombre_curso):
    try:
        res = supabase.table("inscripciones").select("id, alumnos(id, nombre, apellido)").eq("nombre_curso_materia", nombre_curso).not_.is_("alumno_id", "null").execute()
        alumnos = []
        for r in (res.data or []):
            al_raw = r.get('alumnos')
            al = al_raw[0] if isinstance(al_raw, list) and al_raw else al_raw
            if al:
                alumnos.append({'insc_id': r['id'], 'id': al['id'], 'nombre': al['nombre'], 'apellido': al['apellido']})
        return sorted(alumnos, key=lambda x: x['apellido'])
    except: return []

def render_seccion_calendario(sede, es_admin=False):
    cal = get_calendario_sede(sede)
    fecha_inicio = datetime.date.fromisoformat(cal['fecha_inicio']) if cal and cal.get('fecha_inicio') else None
    fecha_fin = datetime.date.fromisoformat(cal['fecha_fin']) if cal and cal.get('fecha_fin') else None
    if fecha_inicio and fecha_fin:
        hoy = datetime.date.today()
        total_dias = (fecha_fin - fecha_inicio).days
        dias_transcurridos = max(0, (hoy - fecha_inicio).days)
        dias_restantes = max(0, (fecha_fin - hoy).days)
        st.markdown(f'''<div class="calendario-box">
            <div class="cal-header">📆 Año Lectivo {fecha_inicio.year}</div>
            <div class="cal-fechas">🟢 Inicio: <b>{fecha_inicio.strftime("%d/%m/%Y")}</b> &nbsp;·&nbsp; 🔴 Fin: <b>{fecha_fin.strftime("%d/%m/%Y")}</b></div>
            <div class="cal-dias">Total: {total_dias} días · Transcurridos: {dias_transcurridos} · Restantes: {dias_restantes}</div>
        </div>''', unsafe_allow_html=True)
    else:
        st.markdown('<div class="calendario-box"><div class="cal-header">📆 Año Lectivo</div><div class="cal-fechas" style="color:#556">Sin fechas configuradas aún.</div></div>', unsafe_allow_html=True)
    if cal:
        url_1c = cal.get('cronograma_1c_url'); nombre_1c = cal.get('cronograma_1c_nombre')
        url_2c = cal.get('cronograma_2c_url'); nombre_2c = cal.get('cronograma_2c_nombre')
        if url_1c: render_cronograma_visor(url_1c, nombre_1c, "Cronograma 1° Cuatrimestre")
        if url_2c: render_cronograma_visor(url_2c, nombre_2c, "Cronograma 2° Cuatrimestre")
    with st.expander("✏️ Editar calendario y cronogramas", expanded=not fecha_inicio):
        with st.form(f"cal_form_{sede}"):
            col1, col2 = st.columns(2)
            fi = col1.date_input("Fecha de inicio:", value=fecha_inicio if fecha_inicio else datetime.date(ANIO_ACTUAL, 3, 1))
            ff = col2.date_input("Fecha de fin:", value=fecha_fin if fecha_fin else datetime.date(ANIO_ACTUAL, 11, 30))
            st.markdown("**📁 Subir cronograma 1° Cuatrimestre**")
            arch_1c = st.file_uploader("Cronograma 1C:", type=['pdf','jpg','jpeg','png','gif','webp','xlsx','xls','docx','doc'], key=f"arch1c_{sede}")
            st.markdown("**📁 Subir cronograma 2° Cuatrimestre**")
            arch_2c = st.file_uploader("Cronograma 2C:", type=['pdf','jpg','jpeg','png','gif','webp','xlsx','xls','docx','doc'], key=f"arch2c_{sede}")
            if st.form_submit_button("💾 Guardar", use_container_width=True):
                if fi >= ff:
                    st.error("La fecha de inicio debe ser anterior a la fecha de fin.")
                else:
                    ok = upsert_calendario_sede(sede, fi, ff)
                    if ok:
                        if arch_1c: subir_cronograma(sede, 1, arch_1c)
                        if arch_2c: subir_cronograma(sede, 2, arch_2c)
                        st.success("✅ Calendario actualizado."); st.rerun()

# =========================================================
# --- FUNCIONES ESTADÍSTICAS (PLOTLY) ---
# =========================================================
COLORES_PLOTLY = ['#4facfe', '#ff4d6d', '#ffc107', '#4ade80', '#a78bfa', '#fb923c', '#38bdf8', '#f472b6']
PLOTLY_LAYOUT = dict(
    paper_bgcolor='#080b10', plot_bgcolor='rgba(255,255,255,0.02)',
    font=dict(family='DM Mono, monospace', color='#e8eaf0', size=12),
    margin=dict(l=20, r=20, t=40, b=20),
    legend=dict(bgcolor='rgba(0,0,0,0)', bordercolor='rgba(255,255,255,0.1)', borderwidth=1),
    xaxis=dict(gridcolor='rgba(255,255,255,0.05)', zerolinecolor='rgba(255,255,255,0.1)'),
    yaxis=dict(gridcolor='rgba(255,255,255,0.05)', zerolinecolor='rgba(255,255,255,0.1)')
)

def get_datos_estadisticas(profesor_id):
    datos = {}
    try:
        res_cursos = supabase.table("inscripciones").select("*").eq("profesor_id", profesor_id).is_("alumno_id", "null").execute()
        for curso in (res_cursos.data or []):
            nombre = curso['nombre_curso_materia']
            nota_ap = curso.get('nota_aprobacion')
            res_al = supabase.table("inscripciones").select("id, alumnos(id, nombre, apellido)").eq("nombre_curso_materia", nombre).not_.is_("alumno_id", "null").execute()
            alumnos = []
            notas = {}
            for r in (res_al.data or []):
                al_raw = r.get('alumnos')
                al = al_raw[0] if isinstance(al_raw, list) and al_raw else al_raw
                if al:
                    alumnos.append({'insc_id': r['id'], 'id': al['id'], 'nombre': al['nombre'], 'apellido': al['apellido']})
                    res_n = supabase.table("notas").select("calificacion, created_at").eq("inscripcion_id", r['id']).order("created_at").execute()
                    notas[r['id']] = [{'val': float(n['calificacion']), 'fecha': n['created_at'][:10]} for n in (res_n.data or [])]
            promedios = {}
            for al in alumnos:
                vals = [x['val'] for x in notas.get(al['insc_id'], [])]
                promedios[al['insc_id']] = round(sum(vals)/len(vals), 2) if vals else None
            datos[nombre] = {'alumnos': alumnos, 'notas': notas, 'promedios': promedios, 'nota_ap': nota_ap}
    except Exception as e:
        st.error(f"Error cargando datos: {e}")
    return datos

def grafico_promedios_por_alumno(nombre_curso, datos_curso):
    alumnos = datos_curso['alumnos']
    promedios = datos_curso['promedios']
    nota_ap = datos_curso.get('nota_ap')
    nombres = [f"{al['apellido'].upper()}, {al['nombre']}" for al in sorted(alumnos, key=lambda x: x['apellido'])]
    vals = [promedios.get(al['insc_id']) for al in sorted(alumnos, key=lambda x: x['apellido'])]
    colores = []
    for v in vals:
        if v is None: colores.append('#556')
        elif nota_ap and v >= float(nota_ap): colores.append('#4facfe')
        else: colores.append('#ff4d6d')
    vals_display = [v if v is not None else 0 for v in vals]
    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=nombres, x=vals_display, orientation='h',
        marker_color=colores,
        text=[str(v) if v is not None else 'S/N' for v in vals],
        textposition='outside',
        hovertemplate='%{y}: %{x}<extra></extra>'
    ))
    if nota_ap:
        fig.add_vline(x=float(nota_ap), line_dash="dash", line_color="#ffc107",
                      annotation_text=f"Aprobación: {nota_ap}", annotation_position="top right",
                      annotation_font_color="#ffc107")
    layout = dict(**PLOTLY_LAYOUT)
    layout['title'] = dict(text="Promedios por Alumno", font=dict(color='#4facfe', size=14))
    layout['xaxis'] = dict(**PLOTLY_LAYOUT['xaxis'], range=[0, 10])
    layout['height'] = max(300, len(nombres) * 40 + 80)
    fig.update_layout(**layout)
    return fig

def grafico_evolucion_notas(nombre_curso, datos_curso):
    alumnos = datos_curso['alumnos']
    notas = datos_curso['notas']
    fig = go.Figure()
    for i, al in enumerate(sorted(alumnos, key=lambda x: x['apellido'])):
        ns = notas.get(al['insc_id'], [])
        if ns:
            color = COLORES_PLOTLY[i % len(COLORES_PLOTLY)]
            fig.add_trace(go.Scatter(
                x=[n['fecha'] for n in ns],
                y=[n['val'] for n in ns],
                mode='lines+markers',
                name=f"{al['apellido'].upper()}, {al['nombre']}",
                line=dict(color=color, width=2),
                marker=dict(size=8, color=color),
                hovertemplate='%{x}: %{y}<extra></extra>'
            ))
    layout = dict(**PLOTLY_LAYOUT)
    layout['title'] = dict(text="Evolución de Notas", font=dict(color='#4facfe', size=14))
    layout['yaxis'] = dict(**PLOTLY_LAYOUT['yaxis'], range=[0, 10])
    layout['height'] = 380
    fig.update_layout(**layout)
    return fig

def grafico_aprobados_desaprobados(nombre_curso, datos_curso):
    alumnos = datos_curso['alumnos']
    promedios = datos_curso['promedios']
    nota_ap = datos_curso.get('nota_ap')
    aprobados = desaprobados = sin_notas = 0
    for al in alumnos:
        p = promedios.get(al['insc_id'])
        if p is None: sin_notas += 1
        elif nota_ap and p >= float(nota_ap): aprobados += 1
        else: desaprobados += 1
    labels = ['Aprobados', 'Desaprobados', 'Sin notas']
    values = [aprobados, desaprobados, sin_notas]
    colors_pie = ['#4facfe', '#ff4d6d', '#445']
    fig = go.Figure(data=[go.Pie(
        labels=labels, values=values,
        hole=0.45,
        marker_colors=colors_pie,
        textinfo='label+percent',
        textfont=dict(color='#e8eaf0', size=12),
        hovertemplate='%{label}: %{value} alumno/s<extra></extra>'
    )])
    layout = dict(**PLOTLY_LAYOUT)
    layout['title'] = dict(text="Aprobados / Desaprobados", font=dict(color='#4facfe', size=14))
    layout['height'] = 350
    layout.pop('xaxis', None); layout.pop('yaxis', None)
    fig.update_layout(**layout)
    return fig

def grafico_comparativa_cursos(datos):
    nombres = []
    promedios_gral = []
    pct_aprobados = []
    for nombre, d in datos.items():
        alumnos = d['alumnos']
        promedios = d['promedios']
        nota_ap = d.get('nota_ap')
        vals = [v for v in promedios.values() if v is not None]
        if not vals: continue
        prom = round(sum(vals)/len(vals), 2)
        aprobados = sum(1 for v in vals if nota_ap and v >= float(nota_ap))
        pct = round(aprobados/len(vals)*100) if vals else 0
        nombre_corto = nombre[:25] + '...' if len(nombre) > 25 else nombre
        nombres.append(nombre_corto)
        promedios_gral.append(prom)
        pct_aprobados.append(pct)
    if not nombres:
        return None
    fig = go.Figure()
    fig.add_trace(go.Bar(
        name='Promedio general', x=nombres, y=promedios_gral,
        marker_color='#4facfe', yaxis='y',
        text=[str(p) for p in promedios_gral], textposition='outside',
        hovertemplate='%{x}<br>Promedio: %{y}<extra></extra>'
    ))
    fig.add_trace(go.Bar(
        name='% Aprobados', x=nombres, y=pct_aprobados,
        marker_color='#4ade80', yaxis='y2',
        text=[f"{p}%" for p in pct_aprobados], textposition='outside',
        hovertemplate='%{x}<br>% Aprobados: %{y}%<extra></extra>'
    ))
    layout = dict(**PLOTLY_LAYOUT)
    layout['title'] = dict(text="Comparativa entre Cursos", font=dict(color='#4facfe', size=14))
    layout['barmode'] = 'group'
    layout['yaxis'] = dict(title='Promedio', range=[0, 11], gridcolor='rgba(255,255,255,0.05)')
    layout['yaxis2'] = dict(title='% Aprobados', range=[0, 110], overlaying='y', side='right', gridcolor='rgba(0,0,0,0)')
    layout['height'] = 400
    fig.update_layout(**layout)
    return fig

def render_tab_estadisticas(profesor_id, mapa_cursos, mapa_cursos_data):
    st.subheader("📈 Estadísticas de Rendimiento")
    if not mapa_cursos:
        no_encontrado("No tenés cursos creados.")
        return
    if not PLOTLY_OK:
        st.error("La librería plotly no está instalada. Agregá 'plotly' a requirements.txt")
        return
    with st.spinner("Cargando datos..."):
        datos = get_datos_estadisticas(profesor_id)
    if not datos:
        no_encontrado("No hay datos suficientes para mostrar estadísticas.")
        return
    cursos_con_notas = {n: d for n, d in datos.items() if any(v is not None for v in d['promedios'].values())}
    if not cursos_con_notas:
        st.info("💡 Aún no hay notas cargadas. Las estadísticas se mostrarán cuando haya calificaciones.")
        return
    sub_est = st.radio("Ver:", ["Por curso", "Comparativa general"], horizontal=True, key="est_sub")
    if sub_est == "Por curso":
        curso_est = st.selectbox("Seleccione Curso:", list(cursos_con_notas.keys()), key="est_curso_sel")
        d = cursos_con_notas[curso_est]
        col1, col2 = st.columns(2)
        with col1:
            st.plotly_chart(grafico_promedios_por_alumno(curso_est, d), use_container_width=True)
        with col2:
            st.plotly_chart(grafico_aprobados_desaprobados(curso_est, d), use_container_width=True)
        tiene_evolucion = any(len(ns) > 1 for ns in d['notas'].values())
        if tiene_evolucion:
            st.plotly_chart(grafico_evolucion_notas(curso_est, d), use_container_width=True)
        else:
            st.caption("💡 La evolución de notas se mostrará cuando haya al menos 2 notas por alumno.")
    else:
        fig_comp = grafico_comparativa_cursos(cursos_con_notas)
        if fig_comp:
            st.plotly_chart(fig_comp, use_container_width=True)
        else:
            no_encontrado("No hay suficientes datos para comparar cursos.")

# =========================================================
# --- FUNCIONES DE BACKUP ---
# =========================================================
def get_ultimo_backup(profesor_id):
    try:
        res = supabase.table("backup_log").select("*").eq("profesor_id", profesor_id).order("fecha", desc=True).limit(1).execute()
        return res.data[0] if res.data else None
    except: return None

def get_historial_backups(profesor_id, limite=10):
    try:
        res = supabase.table("backup_log").select("*").eq("profesor_id", profesor_id).order("fecha", desc=True).limit(limite).execute()
        return res.data or []
    except: return []

def dias_desde_ultimo_backup(profesor_id):
    ub = get_ultimo_backup(profesor_id)
    if not ub: return None
    try:
        fecha_ub = datetime.datetime.fromisoformat(ub['fecha'].replace('Z', '+00:00'))
        ahora = datetime.datetime.now(datetime.timezone.utc)
        return (ahora - fecha_ub).days
    except: return None

def generar_datos_backup(profesor_id, sede):
    """Recolecta todos los datos del profesor para el backup."""
    datos = {
        "meta": {
            "version": "292",
            "sede": sede,
            "fecha_backup": datetime.datetime.now().isoformat(),
        },
        "cursos": [],
        "alumnos": [],
        "notas": [],
        "historial_clases": [],
        "asistencia": [],
        "calendario": None,
    }
    try:
        res_cursos = supabase.table("inscripciones").select("*").eq("profesor_id", profesor_id).is_("alumno_id", "null").execute()
        datos["cursos"] = res_cursos.data or []

        res_insc_al = supabase.table("inscripciones").select("id, alumno_id, nombre_curso_materia, anio_lectivo, nota_aprobacion, hora_inicio, hora_fin, horas_catedra, alumnos(id, nombre, apellido, email)").eq("profesor_id", profesor_id).not_.is_("alumno_id", "null").execute()
        alumnos_vistos = set()
        for r in (res_insc_al.data or []):
            al_raw = r.get('alumnos')
            al = al_raw[0] if isinstance(al_raw, list) and al_raw else al_raw
            if al and al['id'] not in alumnos_vistos:
                datos["alumnos"].append(al)
                alumnos_vistos.add(al['id'])
            insc_clean = {k: v for k, v in r.items() if k != 'alumnos'}
            datos["alumnos"].append(insc_clean) if False else None

        ids_insc = [c['id'] for c in datos["cursos"]]
        for iid in ids_insc:
            res_bit = supabase.table("bitacora").select("*").eq("inscripcion_id", iid).execute()
            datos["historial_clases"].extend(res_bit.data or [])
            res_notas = supabase.table("notas").select("*").eq("inscripcion_id", iid).execute()
            datos["notas"].extend(res_notas.data or [])
            res_asist = supabase.table("asistencia").select("*").eq("inscripcion_id", iid).execute()
            datos["asistencia"].extend(res_asist.data or [])

        res_cal = supabase.table("calendario_sede").select("*").eq("sede", sede).execute()
        datos["calendario"] = res_cal.data[0] if res_cal.data else None

        datos["alumnos"] = []
        for r in (res_insc_al.data or []):
            al_raw = r.get('alumnos')
            al = al_raw[0] if isinstance(al_raw, list) and al_raw else al_raw
            if al:
                datos["alumnos"].append({
                    "id": al['id'], "nombre": al['nombre'],
                    "apellido": al['apellido'], "email": al.get('email', ''),
                    "inscripcion_id": r['id'],
                    "curso": r['nombre_curso_materia'],
                })

    except Exception as e:
        st.error(f"Error generando backup: {e}")
    return datos

def datos_a_json_bytes(datos):
    return json.dumps(datos, ensure_ascii=False, indent=2, default=str).encode('utf-8')

def datos_a_excel_bytes(datos, sede):
    wb = openpyxl.Workbook()
    azul = "1A5276"; azul_claro = "D6EAF8"; gris = "F2F2F2"
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=azul)
    header_align = Alignment(horizontal="center", vertical="center")
    normal_font = Font(name="Calibri", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def crear_hoja(nombre):
        if nombre in wb.sheetnames:
            return wb[nombre]
        return wb.create_sheet(nombre)

    def escribir_headers(ws, headers, col_widths=None):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = border
        if col_widths:
            for col_idx, w in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
        ws.row_dimensions[1].height = 22

    # Hoja Cursos
    ws_c = wb.active; ws_c.title = "Cursos"
    escribir_headers(ws_c, ["Nombre curso", "Horario inicio", "Horario fin", "Nota aprobación", "Hs cátedra", "Bibliografía"], [35, 14, 14, 16, 12, 30])
    for i, c in enumerate(datos.get("cursos", []), 2):
        hi = str(c.get('hora_inicio','') or '')[:5]; hf = str(c.get('hora_fin','') or '')[:5]
        row = [c.get('nombre_curso_materia',''), hi, hf, c.get('nota_aprobacion',''), c.get('horas_catedra',''), c.get('bibliografia','') or '']
        for col_idx, v in enumerate(row, 1):
            cell = ws_c.cell(row=i, column=col_idx, value=v)
            cell.font = normal_font; cell.border = border
            if i % 2 == 0: cell.fill = PatternFill("solid", fgColor=gris)

    # Hoja Alumnos
    ws_a = crear_hoja("Alumnos")
    escribir_headers(ws_a, ["Apellido", "Nombre", "Email", "Curso"], [20, 18, 28, 35])
    for i, al in enumerate(sorted(datos.get("alumnos", []), key=lambda x: x.get('apellido', '')), 2):
        row = [al.get('apellido','').upper(), al.get('nombre',''), al.get('email','') or '', al.get('curso','')]
        for col_idx, v in enumerate(row, 1):
            cell = ws_a.cell(row=i, column=col_idx, value=v)
            cell.font = normal_font; cell.border = border
            if i % 2 == 0: cell.fill = PatternFill("solid", fgColor=gris)

    # Hoja Notas
    ws_n = crear_hoja("Notas")
    escribir_headers(ws_n, ["Inscripción ID", "Alumno ID", "Calificación", "Fecha"], [36, 36, 14, 16])
    for i, n in enumerate(datos.get("notas", []), 2):
        row = [n.get('inscripcion_id',''), n.get('alumno_id',''), n.get('calificacion',''), n.get('created_at','')[:10] if n.get('created_at') else '']
        for col_idx, v in enumerate(row, 1):
            cell = ws_n.cell(row=i, column=col_idx, value=v)
            cell.font = normal_font; cell.border = border
            if i % 2 == 0: cell.fill = PatternFill("solid", fgColor=gris)

    # Hoja Historial
    ws_h = crear_hoja("Historial de Clases")
    escribir_headers(ws_h, ["Inscripción ID", "Fecha", "Contenido", "Suplente", "Tarea 1", "Fecha T1", "Tarea 2", "Fecha T2", "Tarea 3", "Fecha T3"], [36, 12, 40, 20, 25, 12, 25, 12, 25, 12])
    for i, b in enumerate(datos.get("historial_clases", []), 2):
        row = [b.get('inscripcion_id',''), b.get('fecha',''), b.get('contenido_clase','') or '', b.get('profesor_suplente','') or '',
               b.get('tarea1','') or '', b.get('tarea1_fecha','') or '', b.get('tarea2','') or '', b.get('tarea2_fecha','') or '',
               b.get('tarea3','') or '', b.get('tarea3_fecha','') or '']
        for col_idx, v in enumerate(row, 1):
            cell = ws_h.cell(row=i, column=col_idx, value=v)
            cell.font = normal_font; cell.border = border
            if i % 2 == 0: cell.fill = PatternFill("solid", fgColor=gris)

    # Hoja Asistencia
    ws_as = crear_hoja("Asistencia")
    escribir_headers(ws_as, ["Inscripción ID", "Alumno ID", "Fecha", "Estado", "Cuatrimestre", "Hora cátedra"], [36, 36, 12, 12, 14, 14])
    for i, a in enumerate(datos.get("asistencia", []), 2):
        row = [a.get('inscripcion_id',''), a.get('alumno_id',''), a.get('fecha',''), a.get('estado',''), a.get('cuatrimestre','') or '', a.get('hora_catedra','') or '']
        for col_idx, v in enumerate(row, 1):
            cell = ws_as.cell(row=i, column=col_idx, value=v)
            cell.font = normal_font; cell.border = border
            if i % 2 == 0: cell.fill = PatternFill("solid", fgColor=gris)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()

def registrar_backup_log(profesor_id, datos, json_desc=False, excel_desc=False):
    try:
        row = {
            "profesor_id": profesor_id,
            "cant_cursos": len(datos.get("cursos", [])),
            "cant_alumnos": len(datos.get("alumnos", [])),
            "cant_clases": len(datos.get("historial_clases", [])),
            "cant_notas": len(datos.get("notas", [])),
            "cant_asistencia": len(datos.get("asistencia", [])),
            "json_descargado": json_desc,
            "excel_descargado": excel_desc,
        }
        res = supabase.table("backup_log").insert(row).execute()
        return res.data[0]['id'] if res.data else None
    except Exception as e:
        st.error(f"Error registrando backup: {e}")
        return None

def actualizar_backup_log(log_id, json_desc=None, excel_desc=None):
    try:
        upd = {}
        if json_desc is not None: upd["json_descargado"] = json_desc
        if excel_desc is not None: upd["excel_descargado"] = excel_desc
        if upd:
            supabase.table("backup_log").update(upd).eq("id", log_id).execute()
    except: pass

def restaurar_desde_json(profesor_id, datos_json):
    """Restauración total: borra todo y recarga desde el JSON."""
    try:
        res_insc = supabase.table("inscripciones").select("id").eq("profesor_id", profesor_id).execute()
        for r in (res_insc.data or []):
            supabase.table("notas").delete().eq("inscripcion_id", r['id']).execute()
            supabase.table("bitacora").delete().eq("inscripcion_id", r['id']).execute()
            supabase.table("asistencia").delete().eq("inscripcion_id", r['id']).execute()
        supabase.table("inscripciones").delete().eq("profesor_id", profesor_id).execute()

        mapa_alumnos_nuevo = {}
        for al in datos_json.get("alumnos", []):
            alumno_id_orig = al.get('id')
            datos_al = {"nombre": al['nombre'], "apellido": al['apellido']}
            if al.get('email'): datos_al["email"] = al['email']
            res_al = supabase.table("alumnos").insert(datos_al).execute()
            if res_al.data:
                mapa_alumnos_nuevo[alumno_id_orig] = res_al.data[0]['id']

        mapa_insc_nuevo = {}
        for c in datos_json.get("cursos", []):
            insc_id_orig = c.get('id')
            datos_c = {
                "profesor_id": profesor_id,
                "nombre_curso_materia": c.get('nombre_curso_materia'),
                "anio_lectivo": c.get('anio_lectivo'),
                "nota_aprobacion": c.get('nota_aprobacion'),
                "bibliografia": c.get('bibliografia'),
                "hora_inicio": c.get('hora_inicio'),
                "hora_fin": c.get('hora_fin'),
                "horas_catedra": c.get('horas_catedra'),
            }
            res_c = supabase.table("inscripciones").insert(datos_c).execute()
            if res_c.data:
                mapa_insc_nuevo[insc_id_orig] = res_c.data[0]['id']

        for al in datos_json.get("alumnos", []):
            alumno_id_orig = al.get('id')
            insc_id_nuevo = None
            for c in datos_json.get("cursos", []):
                pass
            alumno_nuevo_id = mapa_alumnos_nuevo.get(alumno_id_orig)
            insc_id_orig = al.get('inscripcion_id')
            curso_nombre = al.get('curso')
            insc_id_nuevo = mapa_insc_nuevo.get(insc_id_orig)
            if alumno_nuevo_id and insc_id_nuevo:
                supabase.table("inscripciones").insert({
                    "profesor_id": profesor_id,
                    "alumno_id": alumno_nuevo_id,
                    "nombre_curso_materia": curso_nombre,
                    "anio_lectivo": ANIO_ACTUAL,
                }).execute()

        for n in datos_json.get("notas", []):
            insc_id_nuevo = mapa_insc_nuevo.get(n.get('inscripcion_id'))
            if insc_id_nuevo:
                supabase.table("notas").insert({
                    "inscripcion_id": insc_id_nuevo,
                    "alumno_id": n.get('alumno_id'),
                    "calificacion": n.get('calificacion'),
                }).execute()

        for b in datos_json.get("historial_clases", []):
            insc_id_nuevo = mapa_insc_nuevo.get(b.get('inscripcion_id'))
            if insc_id_nuevo:
                bit = {k: v for k, v in b.items() if k not in ['id', 'inscripcion_id']}
                bit['inscripcion_id'] = insc_id_nuevo
                supabase.table("bitacora").insert(bit).execute()

        for a in datos_json.get("asistencia", []):
            insc_id_nuevo = mapa_insc_nuevo.get(a.get('inscripcion_id'))
            if insc_id_nuevo:
                asist = {k: v for k, v in a.items() if k not in ['id', 'inscripcion_id', 'created_at', 'updated_at']}
                asist['inscripcion_id'] = insc_id_nuevo
                supabase.table("asistencia").insert(asist).execute()

        return True, "Restauración completada exitosamente."
    except Exception as e:
        return False, str(e)

def generar_excel(sede, curso_nombre, curso_data, datos):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumnos y Notas"
    azul = "1A5276"; azul_claro = "D6EAF8"; gris = "F2F2F2"
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=azul)
    header_align = Alignment(horizontal="center", vertical="center")
    subheader_fill = PatternFill("solid", fgColor=azul_claro)
    subheader_font = Font(name="Calibri", bold=True, size=10)
    normal_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells("A1:J1")
    ws["A1"] = f"ClassTrack 360 — Sede: {sede.upper()}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Curso: {curso_nombre}"
    ws["A2"].font = subheader_font
    ws["A2"].fill = subheader_fill
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20
    hi = str(curso_data.get('hora_inicio','') or '')[:5]
    hf = str(curso_data.get('hora_fin','') or '')[:5]
    nota_ap = curso_data.get('nota_aprobacion')
    biblio = curso_data.get('bibliografia','') or ''
    info_txt = f"Horario: {format_horario(hi, hf)}   |   Aprobación: {nota_ap if nota_ap else 'Sin definir'}   |   Generado: {datetime.date.today().strftime('%d/%m/%Y')}"
    ws.merge_cells("A3:J3")
    ws["A3"] = info_txt
    ws["A3"].font = Font(name="Calibri", size=9, color="555555")
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 16
    fila_headers = 5 if not biblio else 6
    if biblio:
        ws.merge_cells("A4:J4")
        ws["A4"] = f"Bibliografía: {biblio}"
        ws["A4"].font = Font(name="Calibri", size=9, italic=True, color="777777")
        ws["A4"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[4].height = 14
    headers = ["#", "Apellido", "Nombre", "Email", "Nota 1", "Nota 2", "Nota 3", "Nota 4", "Nota 5", "Promedio", "Estado"]
    col_widths = [5, 20, 18, 28, 9, 9, 9, 9, 9, 11, 14]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=fila_headers, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[fila_headers].height = 22
    alumnos_data = datos.get('alumnos', [])
    notas_data = datos.get('notas', {})
    for idx, al in enumerate(sorted(alumnos_data, key=lambda x: x['apellido']), 1):
        fila = fila_headers + idx
        ns = notas_data.get(al['insc_id'], [])
        promedio = round(sum(ns)/len(ns), 2) if ns else None
        estado = estado_texto(promedio, nota_ap) if promedio is not None else "Sin notas"
        row_fill = PatternFill("solid", fgColor="FFFFFF") if idx % 2 != 0 else PatternFill("solid", fgColor=gris)
        datos_fila = [idx, al['apellido'].upper(), al['nombre'], al.get('email','') or '',
            ns[0] if len(ns) > 0 else '', ns[1] if len(ns) > 1 else '',
            ns[2] if len(ns) > 2 else '', ns[3] if len(ns) > 3 else '',
            ns[4] if len(ns) > 4 else '', promedio if promedio is not None else '', estado]
        for col_idx, val in enumerate(datos_fila, 1):
            cell = ws.cell(row=fila, column=col_idx, value=val)
            cell.font = normal_font; cell.fill = row_fill; cell.border = border
            if col_idx in [1, 5, 6, 7, 8, 9, 10]: cell.alignment = center_align
            if col_idx == 11 and val == "APROBADO":
                cell.font = Font(name="Calibri", size=10, bold=True, color="1A5276")
            elif col_idx == 11 and val == "DESAPROBADO":
                cell.font = Font(name="Calibri", size=10, bold=True, color="C0392B")
        ws.row_dimensions[fila].height = 18
    ws.freeze_panes = ws.cell(row=fila_headers + 1, column=1)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()

def generar_pdf(sede, curso_nombre, curso_data, incluir_alumnos, incluir_notas, incluir_historial, incluir_resumen, datos):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    azul = colors.HexColor('#1a5276'); gris = colors.HexColor('#555555'); gris_claro = colors.HexColor('#f2f2f2')
    estilo_titulo = ParagraphStyle('titulo', parent=styles['Normal'], fontSize=18, fontName='Helvetica-Bold', textColor=azul, spaceAfter=4, alignment=TA_CENTER)
    estilo_subtitulo = ParagraphStyle('subtitulo', parent=styles['Normal'], fontSize=11, fontName='Helvetica', textColor=gris, spaceAfter=12, alignment=TA_CENTER)
    estilo_seccion = ParagraphStyle('seccion', parent=styles['Normal'], fontSize=13, fontName='Helvetica-Bold', textColor=azul, spaceBefore=16, spaceAfter=8)
    estilo_normal = ParagraphStyle('normal', parent=styles['Normal'], fontSize=9, fontName='Helvetica', spaceAfter=4)
    estilo_small = ParagraphStyle('small', parent=styles['Normal'], fontSize=8, fontName='Helvetica', textColor=gris, spaceAfter=2)
    story = []
    story.append(Paragraph("ClassTrack 360", estilo_titulo))
    story.append(Paragraph(f"Sede: {sede.upper()}  ·  {curso_nombre}", estilo_subtitulo))
    story.append(Paragraph(f"Generado el {datetime.date.today().strftime('%d/%m/%Y')}", estilo_small))
    story.append(HRFlowable(width="100%", thickness=2, color=azul, spaceAfter=16))
    hi = str(curso_data.get('hora_inicio','') or '')[:5]; hf = str(curso_data.get('hora_fin','') or '')[:5]
    nota_ap = curso_data.get('nota_aprobacion'); biblio = curso_data.get('bibliografia','')
    info_curso = []
    if hi and hf: info_curso.append(f"Horario: {hi} -> {hf}")
    if nota_ap: info_curso.append(f"Nota de aprobacion: {nota_ap}")
    if biblio: info_curso.append(f"Bibliografia: {biblio}")
    if info_curso: story.append(Paragraph(" · ".join(info_curso), estilo_small)); story.append(Spacer(1, 8))
    alumnos_data = datos.get('alumnos', []); notas_data = datos.get('notas', {}); historial_data = datos.get('historial', [])
    if incluir_resumen and alumnos_data:
        story.append(Paragraph("Resumen del Curso", estilo_seccion))
        total = len(alumnos_data); aprobados = 0; promedios = []
        for al in alumnos_data:
            ns = notas_data.get(al['insc_id'], [])
            if ns:
                prom = round(sum(ns)/len(ns), 2); promedios.append(prom)
                if nota_ap and prom >= float(nota_ap): aprobados += 1
        prom_gral = round(sum(promedios)/len(promedios), 2) if promedios else "-"
        porc = f"{round(aprobados/total*100)}%" if total > 0 else "-"
        t = Table([["Total alumnos","Con notas","Aprobados","% Aprobados","Promedio general"],
                   [str(total),str(len(promedios)),str(aprobados),porc,str(prom_gral)]], colWidths=[3.2*cm]*5)
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),azul),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('GRID',(0,0),(-1,-1),0.5,colors.grey),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,gris_claro])]))
        story.append(t); story.append(Spacer(1,12))
    if incluir_alumnos and alumnos_data:
        story.append(Paragraph("Listado de Alumnos", estilo_seccion))
        rows = [["#","Apellido","Nombre","Email"]]
        for idx, al in enumerate(sorted(alumnos_data, key=lambda x: x['apellido']), 1):
            rows.append([str(idx), al['apellido'].upper(), al['nombre'], al.get('email','') or '-'])
        t = Table(rows, colWidths=[1*cm,5*cm,5*cm,5*cm])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),azul),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
            ('ALIGN',(0,0),(0,-1),'CENTER'),('GRID',(0,0),(-1,-1),0.5,colors.grey),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,gris_claro])]))
        story.append(t); story.append(Spacer(1,12))
    if incluir_notas and alumnos_data:
        story.append(Paragraph("Notas y Promedios", estilo_seccion))
        rows = [["Alumno","Nota 1","Nota 2","Nota 3","Nota 4","Nota 5","Promedio","Estado"]]
        for al in sorted(alumnos_data, key=lambda x: x['apellido']):
            ns = notas_data.get(al['insc_id'], [])
            fila = [f"{al['apellido'].upper()}, {al['nombre']}"]
            for i in range(5): fila.append(str(ns[i]) if i < len(ns) else "-")
            if ns:
                prom = round(sum(ns)/len(ns), 2)
                fila += [str(prom), estado_texto(prom, nota_ap)]
            else: fila += ["-", "Sin notas"]
            rows.append(fila)
        t = Table(rows, colWidths=[5*cm]+[1.5*cm]*5+[2*cm,2.5*cm])
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),azul),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8),
            ('ALIGN',(1,0),(-1,-1),'CENTER'),('GRID',(0,0),(-1,-1),0.5,colors.grey),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,gris_claro])]))
        story.append(t); story.append(Spacer(1,12))
    if incluir_historial and historial_data:
        story.append(Paragraph("Historial de Clases", estilo_seccion))
        for reg in sorted(historial_data, key=lambda x: x['fecha']):
            fecha_fmt = datetime.date.fromisoformat(reg['fecha']).strftime('%d/%m/%Y')
            suplente = reg.get('profesor_suplente')
            prof = f"Suplente: {suplente}" if suplente else "Titular"
            story.append(Paragraph(f"<b>{fecha_fmt}</b> · {prof}", estilo_normal))
            contenido = reg.get('contenido_clase','') or ''
            if contenido: story.append(Paragraph(f"Contenido: {contenido}", estilo_small))
            for i in range(1,4):
                txt = reg.get(f'tarea{i}'); ft = reg.get(f'tarea{i}_fecha')
                completada = reg.get(f'tarea{i}_completada', False)
                if txt:
                    ft_fmt = datetime.date.fromisoformat(ft).strftime('%d/%m/%Y') if ft else "-"
                    estado_t = " [COMPLETADA]" if completada else ""
                    story.append(Paragraph(f"  Tarea {i}: {txt} (entrega: {ft_fmt}){estado_t}", estilo_small))
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#cccccc'), spaceAfter=6))
    story.append(Spacer(1,20))
    estilo_footer_pdf = ParagraphStyle('footer', parent=styles['Normal'], fontSize=7, fontName='Helvetica', textColor=colors.HexColor('#aaaaaa'), alignment=TA_CENTER)
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#cccccc'), spaceAfter=6))
    story.append(Paragraph("® Sistema diseñado y realizado por Fabián Belledi · 2026", estilo_footer_pdf))
    doc.build(story)
    buffer.seek(0)
    return buffer.read()

def generar_html_impresion(sede, curso_nombre, curso_data, incluir_alumnos, incluir_notas, incluir_historial, incluir_resumen, datos):
    hi = str(curso_data.get('hora_inicio','') or '')[:5]; hf = str(curso_data.get('hora_fin','') or '')[:5]
    nota_ap = curso_data.get('nota_aprobacion'); biblio = curso_data.get('bibliografia','')
    alumnos_data = datos.get('alumnos', []); notas_data = datos.get('notas', {}); historial_data = datos.get('historial', [])
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>ClassTrack 360</title>
    <style>body{{font-family:Arial,sans-serif;font-size:11px;color:#111;margin:20px}}
    h1{{color:#1a5276;font-size:20px}}h2{{color:#1a5276;font-size:14px;border-bottom:2px solid #1a5276;padding-bottom:4px;margin-top:20px}}
    .meta{{color:#666;font-size:10px;margin-bottom:16px}}table{{width:100%;border-collapse:collapse;margin-bottom:16px}}
    th{{background:#1a5276;color:white;padding:6px 8px;font-size:10px}}td{{padding:5px 8px;border-bottom:1px solid #ddd;font-size:10px}}
    tr:nth-child(even){{background:#f5f5f5}}.aprobado{{color:#1a5276;font-weight:bold}}.desaprobado{{color:#c0392b;font-weight:bold}}
    .clase-row{{border-bottom:1px solid #eee;padding:6px 0}}.clase-fecha{{font-weight:bold;color:#1a5276}}
    .clase-tarea{{color:#555;margin-left:12px}}.tarea-done{{color:#aaa;text-decoration:line-through}}
    .footer-doc{{text-align:center;font-size:9px;color:#aaa;margin-top:30px;padding-top:10px;border-top:1px solid #ddd}}
    @media print{{body{{margin:10px}}}}</style></head><body>
    <h1>ClassTrack 360</h1><div class="meta">Sede: <b>{sede.upper()}</b> · Curso: <b>{curso_nombre}</b>"""
    if hi and hf: html += f" · Horario: <b>{hi} → {hf}</b>"
    if nota_ap: html += f" · Aprobación: <b>{nota_ap}</b>"
    html += f"<br>Generado el {datetime.date.today().strftime('%d/%m/%Y')}</div>"
    if biblio: html += f"<p><b>Bibliografía:</b> {biblio}</p>"
    if incluir_resumen and alumnos_data:
        total = len(alumnos_data); aprobados = 0; promedios = []
        for al in alumnos_data:
            ns = notas_data.get(al['insc_id'], [])
            if ns:
                prom = round(sum(ns)/len(ns), 2); promedios.append(prom)
                if nota_ap and prom >= float(nota_ap): aprobados += 1
        prom_gral = round(sum(promedios)/len(promedios), 2) if promedios else "-"
        porc = f"{round(aprobados/total*100)}%" if total > 0 else "-"
        html += f"""<h2>Resumen</h2><table><tr><th>Total</th><th>Con notas</th><th>Aprobados</th><th>%</th><th>Promedio</th></tr>
        <tr><td>{total}</td><td>{len(promedios)}</td><td>{aprobados}</td><td>{porc}</td><td>{prom_gral}</td></tr></table>"""
    if incluir_alumnos and alumnos_data:
        html += "<h2>Alumnos</h2><table><tr><th>#</th><th>Apellido</th><th>Nombre</th><th>Email</th></tr>"
        for idx, al in enumerate(sorted(alumnos_data, key=lambda x: x['apellido']), 1):
            html += f"<tr><td>{idx}</td><td>{al['apellido'].upper()}</td><td>{al['nombre']}</td><td>{al.get('email','') or '-'}</td></tr>"
        html += "</table>"
    if incluir_notas and alumnos_data:
        html += "<h2>Notas</h2><table><tr><th>Alumno</th><th>N1</th><th>N2</th><th>N3</th><th>N4</th><th>N5</th><th>Promedio</th><th>Estado</th></tr>"
        for al in sorted(alumnos_data, key=lambda x: x['apellido']):
            ns = notas_data.get(al['insc_id'], [])
            celdas = "".join([f"<td>{ns[i] if i < len(ns) else '-'}</td>" for i in range(5)])
            if ns:
                prom = round(sum(ns)/len(ns), 2); est = estado_texto(prom, nota_ap)
                clase = "aprobado" if est == "APROBADO" else "desaprobado"
                html += f"<tr><td>{al['apellido'].upper()}, {al['nombre']}</td>{celdas}<td>{prom}</td><td class='{clase}'>{est}</td></tr>"
            else:
                html += f"<tr><td>{al['apellido'].upper()}, {al['nombre']}</td>{celdas}<td>-</td><td>Sin notas</td></tr>"
        html += "</table>"
    if incluir_historial and historial_data:
        html += "<h2>Historial de Clases</h2>"
        for reg in sorted(historial_data, key=lambda x: x['fecha']):
            fecha_fmt = datetime.date.fromisoformat(reg['fecha']).strftime('%d/%m/%Y')
            suplente = reg.get('profesor_suplente'); prof = f"Suplente: {suplente}" if suplente else "Titular"
            html += f"<div class='clase-row'><span class='clase-fecha'>{fecha_fmt}</span> · {prof}"
            contenido = reg.get('contenido_clase','') or ''
            if contenido: html += f"<br><span style='color:#333'>{contenido}</span>"
            for i in range(1,4):
                txt = reg.get(f'tarea{i}'); ft = reg.get(f'tarea{i}_fecha')
                completada = reg.get(f'tarea{i}_completada', False)
                if txt:
                    ft_fmt = datetime.date.fromisoformat(ft).strftime('%d/%m/%Y') if ft else "-"
                    clase_done = " tarea-done" if completada else ""
                    html += f"<div class='clase-tarea{clase_done}'>Tarea {i}: {txt} (entrega: {ft_fmt})</div>"
            html += "</div>"
    html += "<div class='footer-doc'>® Sistema diseñado y realizado por Fabián Belledi · 2026</div>"
    html += "<script>window.onload=function(){window.print();}</script></body></html>"
    return html

def cargar_datos_por_nombre_curso(nombre_curso, inscripcion_id):
    alumnos = []; notas_dict = {}; historial = []
    try:
        res_al = supabase.table("inscripciones").select("id, alumnos(nombre, apellido, email)").eq("nombre_curso_materia", nombre_curso).not_.is_("alumno_id", "null").execute()
        for r in (res_al.data or []):
            al_raw = r.get('alumnos')
            al = al_raw[0] if isinstance(al_raw, list) and al_raw else al_raw
            if al:
                alumnos.append({'insc_id': r['id'], 'nombre': al.get('nombre',''), 'apellido': al.get('apellido',''), 'email': al.get('email','') or ''})
                res_n = supabase.table("notas").select("calificacion").eq("inscripcion_id", r['id']).order("created_at").execute()
                notas_dict[r['id']] = [float(n['calificacion']) for n in (res_n.data or [])]
    except: pass
    try:
        res_h = supabase.table("bitacora").select("*").eq("inscripcion_id", inscripcion_id).order("fecha").execute()
        historial = res_h.data or []
    except: pass
    return {'alumnos': alumnos, 'notas': notas_dict, 'historial': historial}

# ============================================================
# FIN PARTE 1 DE 2 — Continuá pegando desde aquí en Parte 2
# ============================================================
# ============================================================
# INICIO PARTE 2 DE 2 — Pegá esto a continuación de Parte 1
# ============================================================

# =========================================================
# --- LOGIN Y REGISTRO ---
# =========================================================
if st.session_state.user is None:
    _, col, _ = st.columns([1.5, 1, 1.5])
    with col:
        if st.session_state.pantalla_login == 'registro':
            st.markdown("""<div class="login-box">
                <div class="login-logo">Class<span>Track</span> 360</div>
                <div class="login-eyebrow">// Crear cuenta nueva</div>
                <div class="login-title">Registro</div></div>""", unsafe_allow_html=True)
            if st.session_state.get('ok_registro_cuenta'):
                st.success(f"✅ Cuenta creada satisfactoriamente. Ya podés iniciar sesión con la sede '{st.session_state.ok_registro_cuenta}'.")
                st.session_state.ok_registro_cuenta = None
            with st.form("registro", clear_on_submit=False):
                codigo_input = st.text_input("Código de invitación *", placeholder="Ej: AB12CD34")
                sede_input_r = st.text_input("Nombre de sede *", placeholder="Ej: quilmes")
                nombre_input = st.text_input("Tu nombre completo *", placeholder="Ej: García, María")
                clave1 = st.text_input("Contraseña *", type="password")
                clave2 = st.text_input("Repetir contraseña *", type="password")
                submitted_r = st.form_submit_button("✅ CREAR CUENTA", use_container_width=True)
                if submitted_r:
                    errores = []
                    if not codigo_input.strip(): errores.append("El código de invitación es obligatorio.")
                    if not sede_input_r.strip(): errores.append("El nombre de sede es obligatorio.")
                    if not nombre_input.strip(): errores.append("Tu nombre es obligatorio.")
                    if not clave1.strip(): errores.append("La contraseña es obligatoria.")
                    elif clave1 != clave2: errores.append("Las contraseñas no coinciden.")
                    elif len(clave1) < 6: errores.append("La contraseña debe tener al menos 6 caracteres.")
                    if errores:
                        for e in errores: st.error(e)
                    else:
                        try:
                            res_cod = supabase.table("codigos_invitacion").select("*").eq("codigo", codigo_input.strip().upper()).eq("usado", False).execute()
                            if not res_cod.data:
                                st.error("Código inválido o ya utilizado.")
                            else:
                                cod = res_cod.data[0]
                                sede_norm = sede_input_r.strip().lower()
                                res_sede = supabase.table("usuarios").select("id").eq("sede", sede_norm).execute()
                                if res_sede.data:
                                    st.error(f"La sede '{sede_norm}' ya está registrada. Elegí otro nombre.")
                                else:
                                    supabase.table("usuarios").insert({
                                        "sede": sede_norm, "nombre": nombre_input.strip(),
                                        "password_text": clave1, "habilitado": True,
                                        "tipo_cuenta": cod.get('tipo_cuenta', 'permanente'),
                                        "email": f"{sede_norm}@classtrack.com"
                                    }).execute()
                                    supabase.table("codigos_invitacion").update({
                                        "usado": True, "usado_por": sede_norm
                                    }).eq("id", cod['id']).execute()
                                    st.session_state.ok_registro_cuenta = sede_norm
                                    st.session_state.pantalla_login = 'login'
                                    st.rerun()
                        except Exception as e:
                            st.error(f"Error: {e}")
            if st.button("← Volver al inicio de sesión", use_container_width=True):
                st.session_state.pantalla_login = 'login'
                st.rerun()
            st.markdown('<div class="login-footer">© 2026 ClassTrack 360 · v297</div>', unsafe_allow_html=True)
        else:
            st.markdown("""<div class="login-box">
                <div class="login-logo">Class<span>Track</span> 360</div>
                <div class="login-eyebrow">// Acceso al sistema</div>
                <div class="login-title">Iniciar sesión</div></div>""", unsafe_allow_html=True)
            with st.form("login", clear_on_submit=False):
                sede_input = st.text_input("Sede", key="sede_input")
                clave_input = st.text_input("Clave de acceso", type="password", key="clave_input")
                submitted = st.form_submit_button("ENTRAR AL SISTEMA", use_container_width=True)
                if submitted:
                    sede = sede_input.strip().lower()
                    clave = clave_input.strip()
                    try:
                        res = supabase.table("usuarios").select("*").eq("sede", sede).eq("password_text", clave).execute()
                        if res.data:
                            u = res.data[0]
                            if u.get('habilitado') == False:
                                st.error("Tu cuenta está deshabilitada. Contactá al administrador.")
                            else:
                                st.session_state.user = u
                                st.rerun()
                        else:
                            st.error("Sede o clave incorrectos.")
                    except Exception as e:
                        st.error(f"Error de conexión: {e}")
            st.markdown('<div class="registro-link">¿Primera vez? Si tenés un código de invitación podés crear tu cuenta.</div>', unsafe_allow_html=True)
            if st.button("➕ Crear cuenta nueva", use_container_width=True):
                st.session_state.pantalla_login = 'registro'
                st.rerun()
            st.markdown('<div class="login-footer">© 2026 ClassTrack 360 · v297</div>', unsafe_allow_html=True)
    footer()

# =========================================================
# --- PANEL ADMIN ---
# =========================================================
elif st.session_state.user.get('sede', '').lower() == 'admin':
    u_data = st.session_state.user
    with st.sidebar:
        st.markdown('<div class="admin-badge">⚡ Modo Administrador</div>', unsafe_allow_html=True)
        st.write(f"📅 {datetime.date.today().strftime('%d/%m/%Y')}")
        st.markdown("---")
        try:
            res_sedes = supabase.table("usuarios").select("sede, nombre").neq("sede", "admin").execute()
            sedes_disponibles = [s['sede'] for s in res_sedes.data] if res_sedes.data else []
        except:
            sedes_disponibles = []
        if sedes_disponibles:
            opciones_sede = ["— Todas las sedes —"] + sedes_disponibles
            sede_sel = st.selectbox("Ver sede:", opciones_sede, key="admin_sede_sel")
            st.session_state.sede_admin = None if sede_sel == "— Todas las sedes —" else sede_sel
        else:
            st.warning("No hay sedes registradas.")
            sedes_disponibles = []
        st.markdown("---")
        if st.button("🚪 SALIR"):
            st.session_state.user = None
            st.session_state.sede_admin = None
            st.rerun()

    st.markdown('<div class="admin-badge">⚡ Panel de Administración</div>', unsafe_allow_html=True)
    st.title("ClassTrack 360 · Admin")
    footer()

    admin_tabs = st.tabs(["👥 Profesores", "🔑 Códigos de Invitación", "📊 Resumen", "📝 Notas", "📆 Calendarios", "🗑️ Reset de Datos"])

    with admin_tabs[0]:
        st.subheader("👥 Gestión de Profesores")
        try:
            res_profs = supabase.table("usuarios").select("*").neq("sede", "admin").execute()
            profs = res_profs.data or []
            if not profs:
                no_encontrado("No hay profesores registrados.")
            else:
                for prof in profs:
                    habilitado = prof.get('habilitado', True)
                    tipo = prof.get('tipo_cuenta', 'permanente')
                    tag_hab = '<span class="habilitado-tag">✅ HABILITADO</span>' if habilitado else '<span class="deshabilitado-tag">🚫 DESHABILITADO</span>'
                    tag_tipo = f'<span style="color:#aaa;font-size:0.75rem">{tipo.upper()}</span>'
                    st.markdown(f'<div class="planilla-row"><b>{prof.get("sede","").upper()}</b> · {prof.get("nombre","Sin nombre")} &nbsp; {tag_hab} &nbsp; {tag_tipo}</div>', unsafe_allow_html=True)
                    col_h, col_t, col_d = st.columns(3)
                    if habilitado:
                        if col_h.button("🚫 Deshabilitar", key=f"desh_{prof['id']}"):
                            supabase.table("usuarios").update({"habilitado": False}).eq("id", prof['id']).execute()
                            st.success(f"Profesor {prof.get('sede','').upper()} deshabilitado."); st.rerun()
                    else:
                        if col_h.button("✅ Habilitar", key=f"hab_{prof['id']}"):
                            supabase.table("usuarios").update({"habilitado": True}).eq("id", prof['id']).execute()
                            st.success(f"Profesor {prof.get('sede','').upper()} habilitado."); st.rerun()
                    if tipo == 'provisorio':
                        if col_t.button("⬆️ Hacer Permanente", key=f"perm_{prof['id']}"):
                            supabase.table("usuarios").update({"tipo_cuenta": "permanente"}).eq("id", prof['id']).execute()
                            st.success(f"Cuenta de {prof.get('sede','').upper()} cambiada a Permanente."); st.rerun()
                    else:
                        col_t.caption("Cuenta permanente")
                    if col_d.button("🗑️ Eliminar cuenta", key=f"del_prof_{prof['id']}"):
                        st.session_state[f'confirm_del_{prof["id"]}'] = True; st.rerun()
                    if st.session_state.get(f'confirm_del_{prof["id"]}'):
                        st.warning(f"⚠️ ¿Confirmar eliminación de la cuenta {prof.get('sede','').upper()}?")
                        c1, c2 = st.columns(2)
                        if c1.button("✅ Confirmar", key=f"conf_del_{prof['id']}"):
                            supabase.table("usuarios").delete().eq("id", prof['id']).execute()
                            st.session_state[f'confirm_del_{prof["id"]}'] = False
                            st.success("Cuenta eliminada."); st.rerun()
                        if c2.button("❌ Cancelar", key=f"canc_del_{prof['id']}"):
                            st.session_state[f'confirm_del_{prof["id"]}'] = False; st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

    with admin_tabs[1]:
        st.subheader("🔑 Códigos de Invitación")
        col_g1, col_g2 = st.columns([2, 1])
        tipo_nuevo = col_g1.selectbox("Tipo de cuenta:", ["permanente", "provisorio"], key="tipo_cod")
        if col_g2.button("➕ Generar Código", use_container_width=True):
            nuevo_codigo = generar_codigo()
            try:
                supabase.table("codigos_invitacion").insert({"codigo": nuevo_codigo, "tipo_cuenta": tipo_nuevo}).execute()
                st.markdown(f'<div class="codigo-box">🔑 {nuevo_codigo}</div>', unsafe_allow_html=True)
                st.success(f"Código generado. Tipo: {tipo_nuevo.upper()}"); st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")
        st.markdown("---")
        try:
            res_cod = supabase.table("codigos_invitacion").select("*").order("created_at", desc=True).execute()
            codigos = res_cod.data or []
            if not codigos:
                no_encontrado("No hay códigos generados aún.")
            else:
                disponibles = [c for c in codigos if not c.get('usado')]
                usados = [c for c in codigos if c.get('usado')]
                st.caption(f"Disponibles: {len(disponibles)} · Usados: {len(usados)}")
                for cod in codigos:
                    usado = cod.get('usado', False)
                    tipo = cod.get('tipo_cuenta', 'permanente')
                    usado_por = cod.get('usado_por', '')
                    estado_color = "#555" if usado else "#4facfe"
                    estado_txt = f"✅ Usado por: {usado_por}" if usado else "🟢 Disponible"
                    col_a, col_b, col_c = st.columns([2, 2, 1])
                    col_a.markdown(f'<span style="font-family:monospace;font-size:1rem;color:{estado_color};font-weight:700">{cod["codigo"]}</span>', unsafe_allow_html=True)
                    col_b.markdown(f'<span style="font-size:0.8rem;color:#aaa">{tipo.upper()} · {estado_txt}</span>', unsafe_allow_html=True)
                    if not usado:
                        if col_c.button("🗑️", key=f"del_cod_{cod['id']}"):
                            supabase.table("codigos_invitacion").delete().eq("id", cod['id']).execute(); st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")

    with admin_tabs[2]:
        sedes_a_mostrar = sedes_disponibles if not st.session_state.sede_admin else [st.session_state.sede_admin]
        st.subheader("📊 Resumen General" if not st.session_state.sede_admin else f"📊 Resumen — {st.session_state.sede_admin.upper()}")
        for sede_activa in sedes_a_mostrar:
            try:
                res_u = supabase.table("usuarios").select("id, nombre").eq("sede", sede_activa).execute()
                if res_u.data:
                    prof_id = res_u.data[0]['id']
                    res_cursos = supabase.table("inscripciones").select("nombre_curso_materia, hora_inicio, hora_fin").eq("profesor_id", prof_id).is_("alumno_id", "null").execute()
                    res_alumnos = supabase.table("inscripciones").select("id").eq("profesor_id", prof_id).not_.is_("alumno_id", "null").execute()
                    cant_cursos = len(res_cursos.data) if res_cursos.data else 0
                    cant_alumnos = len(res_alumnos.data) if res_alumnos.data else 0
                    with st.expander(f"🏫 {sede_activa.upper()} · {cant_cursos} cursos · {cant_alumnos} alumnos", expanded=len(sedes_a_mostrar) == 1):
                        col1, col2 = st.columns(2)
                        col1.metric("Cursos", cant_cursos)
                        col2.metric("Alumnos", cant_alumnos)
                        if res_cursos.data:
                            for c in res_cursos.data:
                                st.markdown(f'<div class="planilla-row">📖 {c["nombre_curso_materia"]}<br><small style="color:#4facfe;">🕐 {format_horario(c.get("hora_inicio",""), c.get("hora_fin",""))}</small></div>', unsafe_allow_html=True)
                        else:
                            no_encontrado("No hay cursos registrados.")
            except Exception as e:
                st.error(f"Error en sede {sede_activa}: {e}")

    with admin_tabs[3]:
        st.subheader("📝 Notas")
        sedes_a_mostrar_n = sedes_disponibles if not st.session_state.sede_admin else [st.session_state.sede_admin]
        sede_sel_notas = st.selectbox("Sede:", sedes_a_mostrar_n, key="admin_notas_sede") if len(sedes_a_mostrar_n) > 1 else (sedes_a_mostrar_n[0] if sedes_a_mostrar_n else None)
        if sede_sel_notas:
            try:
                res_u = supabase.table("usuarios").select("id").eq("sede", sede_sel_notas).execute()
                if res_u.data:
                    prof_id = res_u.data[0]['id']
                    res_cursos = supabase.table("inscripciones").select("nombre_curso_materia").eq("profesor_id", prof_id).is_("alumno_id", "null").execute()
                    if res_cursos.data:
                        curso_sel = st.selectbox("Curso:", ["---"] + [c['nombre_curso_materia'] for c in res_cursos.data])
                        if curso_sel != "---":
                            res_cur_data = supabase.table("inscripciones").select("nota_aprobacion").eq("nombre_curso_materia", curso_sel).is_("alumno_id", "null").limit(1).execute()
                            nota_ap_admin = res_cur_data.data[0].get('nota_aprobacion') if res_cur_data.data else None
                            res_al = supabase.table("inscripciones").select("id, alumnos(nombre, apellido, email)").eq("nombre_curso_materia", curso_sel).not_.is_("alumno_id", "null").execute()
                            if res_al.data:
                                for r in res_al.data:
                                    al_raw = r.get('alumnos')
                                    al = al_raw[0] if isinstance(al_raw, list) and al_raw else al_raw
                                    if al:
                                        res_notas = supabase.table("notas").select("id, calificacion, created_at").eq("inscripcion_id", r['id']).order("created_at", desc=False).execute()
                                        notas = res_notas.data if res_notas.data else []
                                        filas_html = ""; valores = []
                                        for i, nt in enumerate(notas):
                                            val = float(nt['calificacion']); valores.append(val)
                                            fecha_fmt = datetime.datetime.fromisoformat(nt['created_at'][:10]).strftime('%d/%m/%Y')
                                            filas_html += f'<div class="nota-linea"><span>Nota {i+1} · {fecha_fmt}</span><span class="nota-badge {color_nota(val)}">{val}</span></div>'
                                        if not filas_html:
                                            filas_html = '<div class="nota-linea"><span style="color:#555">Sin notas</span></div>'
                                            promedio_html = estado_html = ""
                                        else:
                                            promedio = round(sum(valores)/len(valores), 2)
                                            promedio_html = f'<div class="promedio-linea"><span>Promedio</span><span class="nota-badge {color_nota(promedio)}">{promedio}</span></div>'
                                            estado_html = estado_aprobacion(promedio, nota_ap_admin)
                                        email_html = f'<div class="email-tag">✉️ {al.get("email","")}</div>' if al.get("email") else ""
                                        st.markdown(f'<div class="alumno-block"><div class="nombre">👤 {al.get("apellido","").upper()}, {al.get("nombre","")} {estado_html if valores else ""}</div>{email_html}{filas_html}{promedio_html if valores else ""}</div>', unsafe_allow_html=True)
                            else:
                                no_encontrado("No hay alumnos inscriptos en este curso.")
                    else:
                        no_encontrado("No hay cursos registrados.")
            except Exception as e:
                st.error(f"Error: {e}")

    with admin_tabs[4]:
        st.subheader("📆 Calendarios por Sede")
        sedes_cal = sedes_disponibles if not st.session_state.sede_admin else [st.session_state.sede_admin]
        if not sedes_cal:
            no_encontrado("No hay sedes registradas.")
        elif st.session_state.sede_admin:
            render_seccion_calendario(st.session_state.sede_admin, es_admin=True)
        else:
            sede_cal_sel = st.selectbox("Seleccionar sede:", sedes_cal, key="admin_cal_sede")
            render_seccion_calendario(sede_cal_sel, es_admin=True)

    with admin_tabs[5]:
        st.subheader("🗑️ Reset de Datos")
        st.markdown('<div class="reset-box"><div class="reset-titulo">⚠️ Zona de Peligro</div>Esta acción borrará TODOS los datos de la sede seleccionada. El usuario NO será eliminado.</div>', unsafe_allow_html=True)
        st.markdown("---")
        if sedes_disponibles:
            sede_reset = st.selectbox("Sede a resetear:", sedes_disponibles, key="reset_sede_sel")
            if st.session_state.confirmar_reset != sede_reset:
                if st.button(f"🔴 BORRAR TODOS LOS DATOS DE {sede_reset.upper()}", type="primary"):
                    st.session_state.confirmar_reset = sede_reset; st.rerun()
            else:
                st.markdown(f'<div class="advertencia-box">⚠️ <b>ADVERTENCIA FINAL</b><br><br>Estás a punto de borrar TODOS los datos de <b>{sede_reset.upper()}</b>. Esta acción es <b>irreversible</b>.</div>', unsafe_allow_html=True)
                col_si, col_no = st.columns(2)
                if col_si.button("✅ SÍ, BORRAR TODO", type="primary", use_container_width=True):
                    ok, msg = reset_completo_sede(sede_reset)
                    if ok:
                        st.success(f"✅ {msg}"); st.session_state.confirmar_reset = None; st.rerun()
                    else:
                        st.error(f"Error: {msg}")
                if col_no.button("❌ NO, CANCELAR", use_container_width=True):
                    st.session_state.confirmar_reset = None; st.rerun()

# =========================================================
# --- APP NORMAL ---
# =========================================================
else:
    u_data = st.session_state.user
    f_hoy = datetime.date.today()
    sede_actual = u_data['sede'].lower()
    es_universitario = es_sistema_universitario(sede_actual)

    total_vencidas_sidebar = get_tareas_vencidas_count(u_data['id'])
    total_clases_sidebar = get_total_clases_sidebar(u_data['id'])

    # =========================================================
    # RECORDATORIO BACKUP AL SALIR
    # =========================================================
    if st.session_state.get('mostrar_salir_backup'):
        ub = get_ultimo_backup(u_data['id'])
        fecha_ub_fmt = ""
        if ub:
            try:
                fecha_ub = datetime.datetime.fromisoformat(ub['fecha'].replace('Z', '+00:00'))
                dias = (datetime.datetime.now(datetime.timezone.utc) - fecha_ub).days
                fecha_ub_fmt = f"Tu último backup fue el **{fecha_ub.strftime('%d/%m/%Y')}** (hace {dias} días)"
                json_st = "✅ JSON descargado" if ub.get('json_descargado') else "❌ JSON no descargado"
                excel_st = "✅ Excel descargado" if ub.get('excel_descargado') else "❌ Excel no descargado"
                fecha_ub_fmt += f" — {json_st} · {excel_st}"
            except: fecha_ub_fmt = "Fecha de último backup no disponible"
        else:
            fecha_ub_fmt = "**Nunca hiciste un backup.**"

        st.markdown(f'''<div class="salir-backup-overlay">
            <div class="salir-backup-titulo">💾 ¿Querés hacer un backup antes de salir?</div>
            <div class="salir-backup-info">{fecha_ub_fmt}</div>
        </div>''', unsafe_allow_html=True)

        col_bk, col_sl = st.columns(2)
        if col_bk.button("💾 Sí, hacer backup primero", use_container_width=True, type="primary"):
            st.session_state.mostrar_salir_backup = False
            st.session_state._ir_a_backup = True
            st.rerun()
        if col_sl.button("🚪 Salir igual", use_container_width=True):
            st.session_state.user = None
            st.session_state.mostrar_salir_backup = False
            st.rerun()
        st.stop()

    with st.sidebar:
        st.header(f"Sede: {u_data['sede'].upper()}")
        tipo_sistema = "🎓 Universitario" if es_universitario else "🏫 Instituto"
        st.caption(tipo_sistema)
        st.write(f"📅 {f_hoy.strftime('%d/%m/%Y')}")
        components.html("""<div style="color:#4facfe;font-family:monospace;font-size:24px;text-align:center;"><div id="c">00:00:00</div></div><script>setInterval(()=>{document.getElementById('c').innerText=new Date().toLocaleTimeString('es-AR',{hour12:false})},1000);</script>""", height=50)
        try:
            res_total = supabase.table("inscripciones").select("id").eq("profesor_id", u_data['id']).eq("anio_lectivo", 2026).not_.is_("alumno_id", "null").execute()
            st.markdown(f'<div class="stat-card">Total Alumnos 2026: <b>{len(res_total.data)}</b></div>', unsafe_allow_html=True)
        except:
            st.markdown('<div class="stat-card">Total Alumnos 2026: <b>-</b></div>', unsafe_allow_html=True)
        if total_clases_sidebar > 0:
            st.markdown(f'<div class="badge-clases">🔢 {total_clases_sidebar} clase{"s" if total_clases_sidebar != 1 else ""} dictada{"s" if total_clases_sidebar != 1 else ""}</div>', unsafe_allow_html=True)
        if total_vencidas_sidebar > 0:
            st.markdown(f'<div class="badge-vencidas">⚠️ {total_vencidas_sidebar} tarea{"s" if total_vencidas_sidebar > 1 else ""} vencida{"s" if total_vencidas_sidebar > 1 else ""}</div>', unsafe_allow_html=True)
        cal_sede = get_calendario_sede(u_data['sede'])
        if cal_sede and cal_sede.get('fecha_inicio') and cal_sede.get('fecha_fin'):
            fi_s = datetime.date.fromisoformat(cal_sede['fecha_inicio'])
            ff_s = datetime.date.fromisoformat(cal_sede['fecha_fin'])
            dias_r = max(0, (ff_s - f_hoy).days)
            st.markdown(f'<div class="stat-card" style="font-size:0.72rem;">📆 Año lectivo<br><b>{fi_s.strftime("%d/%m")}</b> → <b>{ff_s.strftime("%d/%m")}</b><br>{dias_r} días restantes</div>', unsafe_allow_html=True)
        st.markdown("---")
        col_prev, col_next = st.columns(2)
        if col_prev.button("◀", key="cal_prev"):
            if st.session_state.cal_mes == 1: st.session_state.cal_mes = 12; st.session_state.cal_anio -= 1
            else: st.session_state.cal_mes -= 1
            st.rerun()
        if col_next.button("▶", key="cal_next"):
            if st.session_state.cal_mes == 12: st.session_state.cal_mes = 1; st.session_state.cal_anio += 1
            else: st.session_state.cal_mes += 1
            st.rerun()
        st.markdown(render_calendario(st.session_state.cal_mes, st.session_state.cal_anio), unsafe_allow_html=True)
        st.markdown("---")
        # BOTÓN SALIR con chequeo de backup
        if st.button("🚪 SALIR"):
            dias_sin_backup = dias_desde_ultimo_backup(u_data['id'])
            necesita_aviso = dias_sin_backup is None or dias_sin_backup >= 7
            if necesita_aviso:
                st.session_state.mostrar_salir_backup = True
                st.rerun()
            else:
                st.session_state.user = None
                st.rerun()

    footer()

    mapa_cursos = {}
    mapa_cursos_data = {}
    try:
        res_c = supabase.table("inscripciones").select("*").eq("profesor_id", u_data['id']).is_("alumno_id", "null").execute()
        if res_c.data:
            for c in res_c.data:
                if 'nombre_curso_materia' in c:
                    mapa_cursos[c['nombre_curso_materia']] = c['id']
                    mapa_cursos_data[c['nombre_curso_materia']] = c
    except Exception as e:
        st.error(f"Error al cargar cursos: {e}")

    # Índice de tabs — si volvemos de "hacer backup antes de salir"
    tab_default = 10 if st.session_state.get('_ir_a_backup') else 0
    if st.session_state.get('_ir_a_backup'):
        st.session_state._ir_a_backup = False

    tabs = st.tabs([
        "📅 Agenda",
        "📋 Asistencia",
        "📊 Historial de Clases",
        "👥 Alumnos",
        "📝 Notas",
        "📈 Estadísticas",
        "🔢 Contador de Clases",
        "🏗️ Cursos",
        f"📆 Calendario Académico {ANIO_ACTUAL}",
        "🖨️ Impresión",
        "🗄️ Backup",
    ])

    # =========================================================
    # TAB 0 — AGENDA
    # =========================================================
    with tabs[0]:
        if not mapa_cursos:
            no_encontrado("No tenés cursos creados. Andá a la pestaña 🏗️ Cursos para crear uno.")
        else:
            c_ag = st.selectbox("Seleccione Curso:", list(mapa_cursos.keys()), key="ag_sel")
            inscripcion_id = mapa_cursos[c_ag]
            curso_sel_data = mapa_cursos_data.get(c_ag, {})
            hi = str(curso_sel_data.get('hora_inicio', '') or '')[:5]
            hf = str(curso_sel_data.get('hora_fin', '') or '')[:5]
            if hi and hf: st.caption(f"🕐 Horario: {format_horario(hi, hf)}")
            proxima = get_proxima_clase(c_ag, hi, hf)
            if proxima:
                dias_txt = "mañana" if proxima['dias_faltan'] == 1 else f"en {proxima['dias_faltan']} días"
                st.markdown(f'''<div class="proxima-clase-card">
                    <div class="pc-titulo">📅 PRÓXIMA CLASE</div>
                    <div class="pc-fecha">{proxima["fecha_fmt"]}</div>
                    <div class="pc-detalle">🕐 {proxima["horario"]} &nbsp;·&nbsp; {dias_txt}</div>
                </div>''', unsafe_allow_html=True)
            try:
                res_tareas_pend = supabase.table("bitacora").select(
                    "id, fecha, tarea1, tarea1_fecha, tarea1_completada, tarea2, tarea2_fecha, tarea2_completada, tarea3, tarea3_fecha, tarea3_completada"
                ).eq("inscripcion_id", inscripcion_id).execute()
                tareas_vencidas = []
                tareas_pendientes = []
                for reg in (res_tareas_pend.data or []):
                    for i in range(1, 4):
                        txt = reg.get(f'tarea{i}')
                        fecha_t = reg.get(f'tarea{i}_fecha')
                        completada = reg.get(f'tarea{i}_completada', False)
                        if txt and not completada:
                            item = {'bit_id': reg['id'], 'num': i, 'texto': txt, 'fecha': fecha_t, 'clase_fecha': reg['fecha']}
                            if fecha_t and datetime.date.fromisoformat(fecha_t) < f_hoy:
                                tareas_vencidas.append(item)
                            else:
                                tareas_pendientes.append(item)
                if tareas_vencidas:
                    st.markdown(f'<div class="vencidas-header">⚠️ {len(tareas_vencidas)} TAREA{"S" if len(tareas_vencidas) > 1 else ""} VENCIDA{"S" if len(tareas_vencidas) > 1 else ""}</div>', unsafe_allow_html=True)
                    for tp in tareas_vencidas:
                        fecha_fmt = datetime.date.fromisoformat(tp['fecha']).strftime('%d/%m/%Y') if tp['fecha'] else "-"
                        clase_fmt = datetime.date.fromisoformat(tp['clase_fecha']).strftime('%d/%m/%Y') if tp['clase_fecha'] else "-"
                        key_edit = f"{tp['bit_id']}_{tp['num']}"
                        if st.session_state.editando_tarea == key_edit:
                            with st.form(f"edit_tarea_v_{key_edit}"):
                                st.markdown(f"**✏️ Editando Tarea {tp['num']} · Clase del {clase_fmt}**")
                                nuevo_texto = st.text_area("Descripción:", value=tp['texto'], key=f"etv_txt_{key_edit}")
                                nueva_fecha = st.date_input("Nueva fecha de entrega:", value=f_hoy, key=f"etv_fch_{key_edit}")
                                col_g, col_c = st.columns(2)
                                if col_g.form_submit_button("💾 Guardar"):
                                    guardar_edicion_tarea(tp['bit_id'], tp['num'], nuevo_texto, nueva_fecha)
                                if col_c.form_submit_button("❌ Cancelar"):
                                    st.session_state.editando_tarea = None; st.rerun()
                        else:
                            col_t, col_b1, col_b2 = st.columns([5, 1, 1])
                            with col_t:
                                st.markdown(f'''<div class="tarea-card-vencida">
                                    <div class="tarea-titulo">⚠️ VENCIDA · Tarea {tp["num"]} · Clase del {clase_fmt}</div>
                                    <div class="tarea-texto">{tp["texto"]}</div>
                                    <div class="tarea-fecha">📅 Vencida el: {fecha_fmt}</div>
                                </div>''', unsafe_allow_html=True)
                            with col_b1:
                                st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
                                if st.button("✅ Hecha", key=f"comp_v_{tp['bit_id']}_{tp['num']}"):
                                    marcar_tarea(tp['bit_id'], tp['num'], True)
                            with col_b2:
                                st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
                                if st.button("✏️ Editar", key=f"edit_v_{tp['bit_id']}_{tp['num']}"):
                                    st.session_state.editando_tarea = key_edit; st.rerun()
                if tareas_pendientes:
                    st.markdown('<div class="tareas-pendientes-header">📌 TAREAS PENDIENTES:</div>', unsafe_allow_html=True)
                    for tp in tareas_pendientes:
                        fecha_fmt = datetime.date.fromisoformat(tp['fecha']).strftime('%d/%m/%Y') if tp['fecha'] else "-"
                        clase_fmt = datetime.date.fromisoformat(tp['clase_fecha']).strftime('%d/%m/%Y') if tp['clase_fecha'] else "-"
                        key_edit = f"{tp['bit_id']}_{tp['num']}"
                        if st.session_state.editando_tarea == key_edit:
                            with st.form(f"edit_tarea_{key_edit}"):
                                st.markdown(f"**✏️ Editando Tarea {tp['num']} · Clase del {clase_fmt}**")
                                nuevo_texto = st.text_area("Descripción:", value=tp['texto'], key=f"et_txt_{key_edit}")
                                nueva_fecha = st.date_input("Fecha de entrega:", value=datetime.date.fromisoformat(tp['fecha']) if tp['fecha'] else f_hoy, key=f"et_fch_{key_edit}")
                                col_g, col_c = st.columns(2)
                                if col_g.form_submit_button("💾 Guardar"):
                                    guardar_edicion_tarea(tp['bit_id'], tp['num'], nuevo_texto, nueva_fecha)
                                if col_c.form_submit_button("❌ Cancelar"):
                                    st.session_state.editando_tarea = None; st.rerun()
                        else:
                            col_t, col_b1, col_b2 = st.columns([5, 1, 1])
                            with col_t:
                                st.markdown(f'''<div class="tarea-card">
                                    <div class="tarea-titulo">Tarea {tp["num"]} · Clase del {clase_fmt}</div>
                                    <div class="tarea-texto">{tp["texto"]}</div>
                                    <div class="tarea-fecha">📅 Entrega: {fecha_fmt}</div>
                                </div>''', unsafe_allow_html=True)
                            with col_b1:
                                st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
                                if st.button("✅ Hecha", key=f"comp_{tp['bit_id']}_{tp['num']}"):
                                    marcar_tarea(tp['bit_id'], tp['num'], True)
                            with col_b2:
                                st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
                                if st.button("✏️ Editar", key=f"edit_tp_{tp['bit_id']}_{tp['num']}"):
                                    st.session_state.editando_tarea = key_edit; st.rerun()
            except Exception as e:
                st.error(f"Error al cargar tareas: {e}")
            try:
                res_t = supabase.table("bitacora").select("id, tarea_proxima, fecha, tarea_proxima_completada").eq("inscripcion_id", inscripcion_id).not_.is_("tarea_proxima", "null").order("fecha", desc=True).limit(1).execute()
                if res_t.data:
                    reg_legacy = res_t.data[0]
                    tarea_txt = reg_legacy.get('tarea_proxima', '')
                    tarea_fecha = reg_legacy.get('fecha', '')
                    completada_legacy = reg_legacy.get('tarea_proxima_completada', False)
                    if tarea_txt and not completada_legacy:
                        if st.session_state.editando_tarea_legacy == reg_legacy['id']:
                            with st.form(f"edit_legacy_{reg_legacy['id']}"):
                                st.markdown(f"**✏️ Editando tarea de la clase del {tarea_fecha}**")
                                nuevo_txt_legacy = st.text_area("Descripción:", value=tarea_txt, key=f"etl_txt_{reg_legacy['id']}")
                                col_gl, col_cl = st.columns(2)
                                if col_gl.form_submit_button("💾 Guardar"):
                                    guardar_edicion_tarea_legacy(reg_legacy['id'], nuevo_txt_legacy)
                                if col_cl.form_submit_button("❌ Cancelar"):
                                    st.session_state.editando_tarea_legacy = None; st.rerun()
                        else:
                            st.markdown(f'''<div class="tarea-alerta">
                                🔔 TAREA PENDIENTE DE LA CLASE ANTERIOR ({tarea_fecha})<br>
                                <div style="margin-top:10px;border-top:1px solid #ffc107;padding-top:10px;color:#fff;font-weight:400;font-size:1rem;">{tarea_txt}</div>
                            </div>''', unsafe_allow_html=True)
                            col_l1, col_l2 = st.columns(2)
                            if col_l1.button("✅ Marcar como hecha", key=f"legacy_done_{reg_legacy['id']}"):
                                marcar_tarea_proxima(reg_legacy['id'], True)
                            if col_l2.button("✏️ Editar tarea", key=f"legacy_edit_{reg_legacy['id']}"):
                                st.session_state.editando_tarea_legacy = reg_legacy['id']; st.rerun()
            except Exception as e:
                st.error(f"Error al cargar tarea legacy: {e}")
            st.markdown("---")
            st.subheader("📝 Registrar Clase de Hoy")
            # Mensaje de éxito fuera del form
            if st.session_state.get('ok_clase_guardada'):
                st.success("✅ Clase guardada satisfactoriamente.")
                st.session_state.ok_clase_guardada = False
            try:
                res_hoy = supabase.table("bitacora").select("id").eq("inscripcion_id", inscripcion_id).eq("fecha", str(f_hoy)).execute()
                ya_guardado_hoy = len(res_hoy.data) > 0
            except:
                ya_guardado_hoy = False
            if ya_guardado_hoy:
                st.warning("⚠️ Ya existe un registro para HOY en este curso.")
            else:
                col_tit, col_sup = st.columns([3, 1])
                with col_tit:
                    if st.session_state.es_suplente:
                        st.markdown('<span class="suplente-badge">⚠️ Registrando clase como SUPLENTE</span>', unsafe_allow_html=True)
                    else:
                        st.markdown('<span class="titular-badge">👤 Clase dictada por TITULAR</span>', unsafe_allow_html=True)
                with col_sup:
                    if not st.session_state.es_suplente:
                        if st.button("👥 Dictada por suplente", key="btn_suplente"):
                            st.session_state.es_suplente = True; st.rerun()
                    else:
                        if st.button("👤 Volver a titular", key="btn_titular"):
                            st.session_state.es_suplente = False; st.rerun()
                suplente_nombre = ""
                if st.session_state.es_suplente:
                    suplente_nombre = st.text_input("Apellido y Nombre del profesor suplente:", placeholder="Ej: García, María")
                with st.form("f_agenda", clear_on_submit=True):
                    temas = st.text_area("Contenido dictado hoy")
                    st.markdown("---")
                    st.markdown("**📌 Tareas** (podés completar hasta 3, ninguna es obligatoria)")
                    col_t1, col_t2, col_t3 = st.columns(3)
                    with col_t1:
                        st.markdown("**Tarea 1**")
                        tarea1 = st.text_area("Descripción:", key="t1_desc", height=100)
                        fecha1 = st.date_input("Fecha:", key="t1_fecha", value=f_hoy + datetime.timedelta(days=7))
                    with col_t2:
                        st.markdown("**Tarea 2**")
                        tarea2 = st.text_area("Descripción:", key="t2_desc", height=100)
                        fecha2 = st.date_input("Fecha:", key="t2_fecha", value=f_hoy + datetime.timedelta(days=7))
                    with col_t3:
                        st.markdown("**Tarea 3**")
                        tarea3 = st.text_area("Descripción:", key="t3_desc", height=100)
                        fecha3 = st.date_input("Fecha:", key="t3_fecha", value=f_hoy + datetime.timedelta(days=7))
                    if st.form_submit_button("💾 Guardar Clase"):
                        if not temas.strip():
                            st.error("El contenido de la clase no puede estar vacío.")
                        elif st.session_state.es_suplente and not suplente_nombre.strip():
                            st.error("Ingresá el apellido y nombre del profesor suplente.")
                        else:
                            try:
                                supabase.table("bitacora").insert({
                                    "inscripcion_id": inscripcion_id, "fecha": str(f_hoy),
                                    "contenido_clase": temas,
                                    "profesor_suplente": suplente_nombre.strip() if st.session_state.es_suplente else None,
                                    "tarea_proxima": tarea1 or tarea2 or tarea3 or None,
                                    "fecha_tarea": str(fecha1) if tarea1 else (str(fecha2) if tarea2 else (str(fecha3) if tarea3 else None)),
                                    "tarea1": tarea1 or None, "tarea1_fecha": str(fecha1) if tarea1 else None,
                                    "tarea2": tarea2 or None, "tarea2_fecha": str(fecha2) if tarea2 else None,
                                    "tarea3": tarea3 or None, "tarea3_fecha": str(fecha3) if tarea3 else None,
                                    "tarea1_completada": False, "tarea2_completada": False, "tarea3_completada": False,
                                    "tarea_proxima_completada": False,
                                }).execute()
                                st.session_state.es_suplente = False
                                st.session_state.ok_clase_guardada = True
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error: {e}")

    # =========================================================
    # TAB 1 — ASISTENCIA
    # =========================================================
    with tabs[1]:
        st.subheader("📋 Asistencia")
        if not mapa_cursos:
            no_encontrado("No tenés cursos creados.")
        else:
            sub_asist = st.radio("Acción:", ["📝 Tomar Lista", "📊 Ver Resumen"], horizontal=True, key="asist_sub")
            if sub_asist == "📝 Tomar Lista":
                curso_asist = st.selectbox("Seleccione Curso:", list(mapa_cursos.keys()), key="asist_curso_sel")
                inscripcion_id_asist = mapa_cursos[curso_asist]
                curso_data_asist = mapa_cursos_data.get(curso_asist, {})
                fecha_asist = st.date_input("Fecha de la clase:", value=f_hoy, key="asist_fecha")
                alumnos_asist = get_alumnos_curso(curso_asist)
                if not alumnos_asist:
                    no_encontrado("No hay alumnos inscriptos en este curso.")
                else:
                    asist_existente = get_asistencia_fecha(inscripcion_id_asist, fecha_asist)
                    if es_universitario:
                        horas_catedra = int(curso_data_asist.get('horas_catedra') or 1)
                        cuatrimestre_asist = get_cuatrimestre_actual()
                        st.caption(f"🎓 Sistema Universitario · {horas_catedra} hs cátedra · Cuatrimestre {cuatrimestre_asist}")
                        estados_existentes = {}
                        for r in asist_existente:
                            h = r.get('hora_catedra', 1) or 1
                            if h not in estados_existentes: estados_existentes[h] = {}
                            estados_existentes[h][r['alumno_id']] = r['estado']
                        estados_por_hora = {}
                        for hora in range(1, horas_catedra + 1):
                            st.markdown(f"**🕐 Hora cátedra {hora}**")
                            estados_hora = {}
                            for al in alumnos_asist:
                                estado_prev = estados_existentes.get(hora, {}).get(al['id'], 'presente')
                                col_n, col_e = st.columns([3, 2])
                                col_n.markdown(f'<div class="asist-nombre">👤 {al["apellido"].upper()}, {al["nombre"]}</div>', unsafe_allow_html=True)
                                estado_sel = col_e.selectbox("Estado:", ["presente", "tarde", "ausente"],
                                    index=["presente", "tarde", "ausente"].index(estado_prev),
                                    key=f"asist_u_{hora}_{al['id']}", label_visibility="collapsed")
                                estados_hora[al['id']] = estado_sel
                            estados_por_hora[hora] = estados_hora
                            if hora < horas_catedra: st.markdown("---")
                        if st.button("💾 Guardar Asistencia", type="primary", use_container_width=True, key="btn_guardar_asist_u"):
                            ok = guardar_asistencia_universitario(inscripcion_id_asist, fecha_asist, cuatrimestre_asist, horas_catedra, estados_por_hora)
                            if ok: st.success("✅ Asistencia guardada correctamente."); st.rerun()
                    else:
                        st.caption(f"🏫 Sistema Instituto · {fecha_asist.strftime('%A %d/%m/%Y').capitalize()}")
                        estados_existentes = {r['alumno_id']: r['estado'] for r in asist_existente}
                        estados_nuevos = {}
                        for al in alumnos_asist:
                            estado_prev = estados_existentes.get(al['id'], 'presente')
                            col_n, col_e = st.columns([3, 2])
                            col_n.markdown(f'<div class="asist-nombre">👤 {al["apellido"].upper()}, {al["nombre"]}</div>', unsafe_allow_html=True)
                            estado_sel = col_e.selectbox("Estado:", ["presente", "tarde", "ausente"],
                                index=["presente", "tarde", "ausente"].index(estado_prev),
                                key=f"asist_i_{al['id']}", label_visibility="collapsed")
                            estados_nuevos[al['id']] = estado_sel
                        if st.button("💾 Guardar Asistencia", type="primary", use_container_width=True, key="btn_guardar_asist_i"):
                            ok = guardar_asistencia_instituto(inscripcion_id_asist, fecha_asist, estados_nuevos)
                            if ok: st.success("✅ Asistencia guardada correctamente."); st.rerun()
                    if asist_existente:
                        st.caption("ℹ️ Ya hay asistencia registrada para esta fecha. Guardar reemplazará los datos existentes.")
            else:
                # --- BUSCADOR POR ALUMNO ---
                busq_al_asist = st.text_input("🔍 Buscar alumno por nombre o apellido:", key="busq_al_asist", placeholder="Escribí para buscar...")
                if busq_al_asist.strip():
                    # Buscar en todos los alumnos del profesor
                    try:
                        res_todos = supabase.table("inscripciones").select("alumno_id, alumnos(id, nombre, apellido)").eq("profesor_id", u_data['id']).not_.is_("alumno_id", "null").execute()
                        alumnos_encontrados = {}
                        for r in (res_todos.data or []):
                            al_raw = r.get('alumnos')
                            al = al_raw[0] if isinstance(al_raw, list) and al_raw else al_raw
                            if al and al['id'] not in alumnos_encontrados:
                                nombre_completo = f"{al.get('apellido','')} {al.get('nombre','')}".lower()
                                if busq_al_asist.lower() in nombre_completo:
                                    alumnos_encontrados[al['id']] = al
                        if not alumnos_encontrados:
                            no_encontrado(f"No se encontró ningún alumno con '{busq_al_asist}'.")
                        else:
                            anio_busq = st.selectbox("Año lectivo:", [ANIO_ACTUAL, ANIO_ACTUAL - 1], key="asist_busq_anio")
                            for al_id, al in alumnos_encontrados.items():
                                registros = get_asistencia_anual_alumno(al_id, u_data['id'], anio_busq)
                                total_p = sum(1 for r in registros if r['estado'] == 'presente')
                                total_t = sum(1 for r in registros if r['estado'] == 'tarde')
                                total_a = sum(1 for r in registros if r['estado'] == 'ausente')
                                dias_html = ""
                                for r in registros:
                                    fecha_fmt = datetime.date.fromisoformat(r['fecha']).strftime('%d/%m/%Y')
                                    curso_corto = r['curso'][:30] + '...' if len(r['curso']) > 30 else r['curso']
                                    hora_label = f" · Hs {r['hora_catedra']}" if r.get('hora_catedra') else ""
                                    if r['estado'] == 'presente':
                                        dias_html += f'<div class="resumen-fila"><span>✅ {fecha_fmt} <span style="color:#556;font-size:0.75rem">· {curso_corto}{hora_label}</span></span><span class="asist-presente">PRESENTE</span></div>'
                                    elif r['estado'] == 'tarde':
                                        dias_html += f'<div class="resumen-fila"><span>🕐 {fecha_fmt} <span style="color:#556;font-size:0.75rem">· {curso_corto}{hora_label}</span></span><span class="asist-tarde">TARDE</span></div>'
                                    elif r['estado'] == 'ausente':
                                        dias_html += f'<div class="resumen-fila"><span>❌ {fecha_fmt} <span style="color:#556;font-size:0.75rem">· {curso_corto}{hora_label}</span></span><span class="asist-ausente">AUSENTE</span></div>'
                                if not dias_html:
                                    dias_html = '<div class="resumen-fila"><span style="color:#556">Sin registros para este año.</span></div>'
                                st.markdown(f'''<div class="resumen-asist">
                                    <div class="resumen-asist-titulo">👤 {al.get("apellido","").upper()}, {al.get("nombre","")} · Año {anio_busq}</div>
                                    {dias_html}
                                    <div style="margin-top:10px;padding-top:8px;border-top:1px solid rgba(79,172,254,0.2);">
                                        <div class="resumen-fila"><span>✅ Total Presentes</span><span style="color:#4facfe;font-weight:700">{total_p} {"hs" if es_universitario else ""}</span></div>
                                        <div class="resumen-fila"><span>🕐 Total Tardes</span><span style="color:#ffc107;font-weight:700">{total_t} {"hs" if es_universitario else ""}</span></div>
                                        <div class="resumen-fila"><span>❌ Total Ausentes</span><span style="color:#ff4d6d;font-weight:700">{total_a} {"hs" if es_universitario else ""}</span></div>
                                    </div>
                                </div>''', unsafe_allow_html=True)
                    except Exception as e:
                        st.error(f"Error: {e}")
                    st.markdown("---")

                curso_res = st.selectbox("Seleccione Curso:", list(mapa_cursos.keys()), key="asist_res_curso")
                inscripcion_id_res = mapa_cursos[curso_res]
                alumnos_res = get_alumnos_curso(curso_res)
                if not alumnos_res:
                    no_encontrado("No hay alumnos inscriptos en este curso.")
                else:
                    if es_universitario:
                        st.caption("🎓 Resumen por cuatrimestre")
                        col_c1, col_c2 = st.columns(2)
                        for cuatri in [1, 2]:
                            resumen = get_resumen_asistencia_universitario(inscripcion_id_res, cuatri)
                            with (col_c1 if cuatri == 1 else col_c2):
                                st.markdown(f"**{cuatri}° Cuatrimestre**")
                                if not resumen:
                                    no_encontrado("Sin datos aún.")
                                else:
                                    for al in alumnos_res:
                                        datos_al = resumen.get(al['id'], {'presente': 0, 'tarde': 0, 'ausente': 0})
                                        p = datos_al['presente']; t = datos_al['tarde']; a = datos_al['ausente']
                                        total_hs = p + t + a
                                        st.markdown(f'''<div class="resumen-asist">
                                            <div class="resumen-asist-titulo">👤 {al["apellido"].upper()}, {al["nombre"]}</div>
                                            <div class="resumen-fila"><span>✅ Presente</span><span>{p} hs</span></div>
                                            <div class="resumen-fila"><span>🕐 Tarde</span><span>{t} hs</span></div>
                                            <div class="resumen-fila"><span>❌ Ausente</span><span>{a} hs</span></div>
                                            <div class="resumen-fila" style="color:#4facfe;font-weight:700"><span>Total hs</span><span>{total_hs}</span></div>
                                        </div>''', unsafe_allow_html=True)
                    else:
                        col_m1, col_m2 = st.columns(2)
                        mes_res = col_m1.selectbox("Mes:", list(range(1, 13)), format_func=lambda x: MESES_ES[x-1], index=f_hoy.month - 1, key="asist_mes_res")
                        anio_res = col_m2.selectbox("Año:", [ANIO_ACTUAL, ANIO_ACTUAL - 1], key="asist_anio_res")
                        resumen = get_resumen_asistencia_instituto(inscripcion_id_res, mes_res, anio_res)
                        if not resumen:
                            no_encontrado(f"Sin registros para {MESES_ES[mes_res-1]} {anio_res}.")
                        else:
                            fechas_mes = sorted(set(fecha for aid_data in resumen.values() for fecha in aid_data.keys()))
                            for al in alumnos_res:
                                datos_al = resumen.get(al['id'], {})
                                presentes = sum(1 for e in datos_al.values() if e == 'presente')
                                tardes = sum(1 for e in datos_al.values() if e == 'tarde')
                                ausentes = sum(1 for e in datos_al.values() if e == 'ausente')
                                dias_html = ""
                                for fecha in fechas_mes:
                                    estado_d = datos_al.get(fecha, '')
                                    fecha_fmt = datetime.date.fromisoformat(fecha).strftime('%d/%m/%Y')
                                    if estado_d == 'presente': dias_html += f'<span class="asist-presente">{fecha_fmt}</span> '
                                    elif estado_d == 'tarde': dias_html += f'<span class="asist-tarde">{fecha_fmt} T</span> '
                                    elif estado_d == 'ausente': dias_html += f'<span class="asist-ausente">{fecha_fmt} A</span> '
                                st.markdown(f'''<div class="resumen-asist">
                                    <div class="resumen-asist-titulo">👤 {al["apellido"].upper()}, {al["nombre"]}</div>
                                    <div style="margin-bottom:8px;font-size:0.82rem">{dias_html}</div>
                                    <div class="resumen-fila"><span>✅ Presentes</span><span>{presentes}</span></div>
                                    <div class="resumen-fila"><span>🕐 Tardes</span><span>{tardes}</span></div>
                                    <div class="resumen-fila"><span>❌ Ausentes</span><span>{ausentes}</span></div>
                                </div>''', unsafe_allow_html=True)

    # =========================================================
    # TAB 2 — HISTORIAL DE CLASES
    # =========================================================
    with tabs[2]:
        st.subheader("📋 Historial de Clases")
        if not mapa_cursos:
            no_encontrado("No tenés cursos creados.")
        else:
            st.markdown('<div class="filtros-box">', unsafe_allow_html=True)
            col_f1, col_f2, col_f3 = st.columns(3)
            hist_curso = col_f1.selectbox("Curso:", ["Todos"] + list(mapa_cursos.keys()), key="hist_curso_sel")
            hist_tipo_prof = col_f2.selectbox("Profesor:", ["Todos", "👤 Titular", "👥 Suplente"], key="hist_tipo_prof_sel")
            hist_contenido = col_f3.text_input("🔍 Buscar en contenido:", key="hist_contenido_sel")
            col_f4, col_f5, col_f6 = st.columns(3)
            hist_anio = col_f4.selectbox("Año lectivo:", list(range(ANIO_ACTUAL, ANIO_ACTUAL - 5, -1)), key="hist_anio_sel")
            hist_desde = col_f5.date_input("Desde:", value=datetime.date(hist_anio, 1, 1), key="hist_desde_sel")
            hist_hasta = col_f6.date_input("Hasta:", value=datetime.date(hist_anio, 12, 31), key="hist_hasta_sel")
            st.markdown('</div>', unsafe_allow_html=True)
            try:
                ids_insc = list(mapa_cursos.values()) if hist_curso == "Todos" else [mapa_cursos[hist_curso]]
                todos_registros = []
                for iid in ids_insc:
                    nombre_curso_hist = [k for k, v in mapa_cursos.items() if v == iid][0]
                    res_b = supabase.table("bitacora").select("*").eq("inscripcion_id", iid).gte("fecha", str(hist_desde)).lte("fecha", str(hist_hasta)).order("fecha", desc=True).execute()
                    for reg in (res_b.data or []):
                        reg['_curso'] = nombre_curso_hist
                        todos_registros.append(reg)
                filtrados = []
                for reg in todos_registros:
                    suplente = reg.get('profesor_suplente')
                    if hist_tipo_prof == "👤 Titular" and suplente: continue
                    if hist_tipo_prof == "👥 Suplente" and not suplente: continue
                    if hist_contenido.strip() and hist_contenido.lower() not in (reg.get('contenido_clase', '') or '').lower(): continue
                    filtrados.append(reg)
                filtrados.sort(key=lambda x: x['fecha'], reverse=True)
                if not filtrados:
                    no_encontrado("No se encontraron clases con los filtros seleccionados.")
                else:
                    st.caption(f"📊 {len(filtrados)} clase/s encontrada/s")
                    if st.session_state.get('ok_bitacora_editada'):
                        st.success("✅ Registro actualizado satisfactoriamente.")
                        st.session_state.ok_bitacora_editada = False
                    for reg in filtrados:
                        suplente_h = reg.get('profesor_suplente')
                        prof_label = f"👥 Suplente: {suplente_h}" if suplente_h else "👤 Titular"
                        fecha_fmt = datetime.date.fromisoformat(reg['fecha']).strftime('%d/%m/%Y')
                        with st.expander(f"📅 {fecha_fmt} · {reg['_curso']} · {prof_label}"):
                            if suplente_h:
                                st.markdown(f'<span class="suplente-badge">👤 Clase dictada por suplente: {suplente_h}</span>', unsafe_allow_html=True)
                            else:
                                st.markdown('<span class="titular-badge">👤 Clase dictada por titular</span>', unsafe_allow_html=True)
                            if st.session_state.editando_bitacora == reg['id']:
                                with st.form(f"edit_bit_h_{reg['id']}"):
                                    t_edit = st.text_area("Contenido dictado:", value=reg.get('contenido_clase', ''))
                                    es_sup_edit = st.checkbox("¿Clase dictada por suplente?", value=bool(suplente_h), key=f"sup_chk_h_{reg['id']}")
                                    sup_nombre_edit = ""
                                    if es_sup_edit:
                                        sup_nombre_edit = st.text_input("Apellido y Nombre del suplente:", value=suplente_h or "", key=f"sup_nom_h_{reg['id']}")
                                    st.markdown("**Tareas:**")
                                    col_t1, col_t2, col_t3 = st.columns(3)
                                    with col_t1:
                                        st.markdown("**Tarea 1**")
                                        t1_e = st.text_area("Descripción:", value=reg.get('tarea1','') or '', key=f"et1_h_{reg['id']}")
                                        f1_e = st.date_input("Fecha:", value=datetime.date.fromisoformat(reg['tarea1_fecha']) if reg.get('tarea1_fecha') else f_hoy, key=f"ef1_h_{reg['id']}")
                                    with col_t2:
                                        st.markdown("**Tarea 2**")
                                        t2_e = st.text_area("Descripción:", value=reg.get('tarea2','') or '', key=f"et2_h_{reg['id']}")
                                        f2_e = st.date_input("Fecha:", value=datetime.date.fromisoformat(reg['tarea2_fecha']) if reg.get('tarea2_fecha') else f_hoy, key=f"ef2_h_{reg['id']}")
                                    with col_t3:
                                        st.markdown("**Tarea 3**")
                                        t3_e = st.text_area("Descripción:", value=reg.get('tarea3','') or '', key=f"et3_h_{reg['id']}")
                                        f3_e = st.date_input("Fecha:", value=datetime.date.fromisoformat(reg['tarea3_fecha']) if reg.get('tarea3_fecha') else f_hoy, key=f"ef3_h_{reg['id']}")
                                    col_e1, col_e2 = st.columns(2)
                                    if col_e1.form_submit_button("💾 Guardar Cambios"):
                                        try:
                                            supabase.table("bitacora").update({
                                                "contenido_clase": t_edit,
                                                "profesor_suplente": sup_nombre_edit.strip() if es_sup_edit and sup_nombre_edit.strip() else None,
                                                "tarea1": t1_e or None, "tarea1_fecha": str(f1_e) if t1_e else None,
                                                "tarea2": t2_e or None, "tarea2_fecha": str(f2_e) if t2_e else None,
                                                "tarea3": t3_e or None, "tarea3_fecha": str(f3_e) if t3_e else None,
                                            }).eq("id", reg['id']).execute()
                                            st.session_state.editando_bitacora = None
                                            st.session_state.ok_bitacora_editada = True
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"Error: {e}")
                                    if col_e2.form_submit_button("❌ Cancelar"):
                                        st.session_state.editando_bitacora = None; st.rerun()
                            else:
                                st.write(f"**Contenido:** {reg.get('contenido_clase', '-')}")
                                for i in range(1, 4):
                                    txt = reg.get(f'tarea{i}'); ft = reg.get(f'tarea{i}_fecha')
                                    completada = reg.get(f'tarea{i}_completada', False)
                                    if txt:
                                        vencida = ft and datetime.date.fromisoformat(ft) < f_hoy and not completada
                                        if completada:
                                            st.markdown(f'''<div class="tarea-card-done"><div class="tarea-titulo">✅ Tarea {i} — COMPLETADA</div><div class="tarea-texto">{txt}</div><div class="tarea-fecha">📅 {datetime.date.fromisoformat(ft).strftime("%d/%m/%Y") if ft else "-"}</div></div>''', unsafe_allow_html=True)
                                            if st.button("↩️ Desmarcar", key=f"descomp_h_{reg['id']}_{i}"):
                                                marcar_tarea(reg['id'], i, False)
                                        elif vencida:
                                            st.markdown(f'''<div class="tarea-card-vencida"><div class="tarea-titulo">⚠️ VENCIDA · Tarea {i}</div><div class="tarea-texto">{txt}</div><div class="tarea-fecha">📅 Venció el: {datetime.date.fromisoformat(ft).strftime("%d/%m/%Y")}</div></div>''', unsafe_allow_html=True)
                                            if st.button(f"✅ Marcar hecha", key=f"comp_vh_{reg['id']}_{i}"):
                                                marcar_tarea(reg['id'], i, True)
                                        else:
                                            st.markdown(f'''<div class="tarea-card"><div class="tarea-titulo">Tarea {i}</div><div class="tarea-texto">{txt}</div><div class="tarea-fecha">📅 {datetime.date.fromisoformat(ft).strftime("%d/%m/%Y") if ft else "-"}</div></div>''', unsafe_allow_html=True)
                                col_b1, col_b2 = st.columns([1, 5])
                                if col_b1.button("✏️ Editar", key=f"edit_b_h_{reg['id']}"):
                                    st.session_state.editando_bitacora = reg['id']; st.rerun()
                                if col_b2.button("🗑️ Borrar", key=f"del_b_h_{reg['id']}"):
                                    try:
                                        supabase.table("bitacora").delete().eq("id", reg['id']).execute()
                                        st.success("Registro eliminado."); st.rerun()
                                    except Exception as e:
                                        st.error(f"Error: {e}")
            except Exception as e:
                st.error(f"Error en historial: {e}")

    # =========================================================
    # TAB 3 — ALUMNOS
    # =========================================================
    with tabs[3]:
        sub_al = st.radio("Acción:", ["Ver Lista", "Registrar Alumno Nuevo"], horizontal=True)
        if sub_al == "Registrar Alumno Nuevo":
            if not mapa_cursos:
                no_encontrado("Primero creá un curso en la pestaña 🏗️ Cursos.")
            else:
                if st.session_state.get('ok_alumno_registrado'):
                    st.success(f"✅ Alumno {st.session_state.ok_alumno_registrado} registrado satisfactoriamente.")
                    st.session_state.ok_alumno_registrado = None
                with st.form("new_al", clear_on_submit=True):
                    n = st.text_input("Nombre *")
                    a = st.text_input("Apellido *")
                    e = st.text_input("Email (opcional)", placeholder="alumno@mail.com")
                    c_sel = st.selectbox("Asignar a curso:", list(mapa_cursos.keys()))
                    if st.form_submit_button("💾 REGISTRAR"):
                        if not n.strip() or not a.strip():
                            st.error("Nombre y Apellido son obligatorios.")
                        else:
                            try:
                                datos_alumno = {"nombre": n.strip(), "apellido": a.strip()}
                                if e.strip(): datos_alumno["email"] = e.strip()
                                ra = supabase.table("alumnos").insert(datos_alumno).execute()
                                if ra.data:
                                    supabase.table("inscripciones").insert({"alumno_id": ra.data[0]['id'], "profesor_id": u_data['id'], "nombre_curso_materia": c_sel, "anio_lectivo": 2026}).execute()
                                    st.session_state.ok_alumno_registrado = f"{a.upper()}, {n}"
                                    st.rerun()
                            except Exception as e_err:
                                st.error(f"Error: {e_err}")
        else:
            if not mapa_cursos:
                no_encontrado("No hay cursos creados.")
            else:
                c_v = st.selectbox("Filtrar por curso:", ["---"] + list(mapa_cursos.keys()), key="curso_alumnos_sel")
                busqueda = st.text_input("🔍 Buscar alumno por nombre o apellido:", key="busq_alumno_input")
                if c_v != "---":
                    try:
                        res_al = supabase.table("inscripciones").select("id, alumnos(id, nombre, apellido, email)").eq("nombre_curso_materia", c_v).not_.is_("alumno_id", "null").execute()
                        alumnos_filtrados = []
                        for r in res_al.data:
                            al_raw = r.get('alumnos')
                            al = al_raw[0] if isinstance(al_raw, list) and al_raw else al_raw
                            if al:
                                if busqueda.strip() == "" or busqueda.lower() in al.get('nombre','').lower() or busqueda.lower() in al.get('apellido','').lower():
                                    alumnos_filtrados.append((r, al))
                        if not res_al.data: no_encontrado("No hay alumnos inscriptos en este curso.")
                        elif not alumnos_filtrados: no_encontrado(f"No se encontró ningún alumno con '{busqueda}'.")
                        else:
                            st.info(f"Cantidad de alumnos: {len(alumnos_filtrados)}")
                            if st.session_state.get('ok_alumno_editado'):
                                st.success("✅ Alumno actualizado satisfactoriamente.")
                                st.session_state.ok_alumno_editado = False
                            for r, al in alumnos_filtrados:
                                if st.session_state.editando_alumno == r['id']:
                                    with st.form(f"edit_al_{r['id']}"):
                                        n_edit = st.text_input("Nombre:", value=al.get('nombre', ''))
                                        a_edit = st.text_input("Apellido:", value=al.get('apellido', ''))
                                        e_edit = st.text_input("Email (opcional):", value=al.get('email', '') or '')
                                        col_a1, col_a2 = st.columns(2)
                                        if col_a1.form_submit_button("💾 Guardar"):
                                            try:
                                                supabase.table("alumnos").update({"nombre": n_edit.strip(), "apellido": a_edit.strip(), "email": e_edit.strip() if e_edit.strip() else None}).eq("id", al['id']).execute()
                                                st.session_state.editando_alumno = None
                                                st.session_state.ok_alumno_editado = True
                                                st.rerun()
                                            except Exception as e_err:
                                                st.error(f"Error: {e_err}")
                                        if col_a2.form_submit_button("❌ Cancelar"):
                                            st.session_state.editando_alumno = None; st.rerun()
                                else:
                                    email_display = f'<br><span class="email-tag">✉️ {al.get("email","")}</span>' if al.get('email') else ''
                                    st.markdown(f'<div class="planilla-row">👤 {al.get("apellido","").upper()}, {al.get("nombre","")}{email_display}</div>', unsafe_allow_html=True)
                                    ab1, ab2 = st.columns(2)
                                    if ab1.button("✏️ Editar", key=f"eal_{r['id']}"):
                                        st.session_state.editando_alumno = r['id']; st.rerun()
                                    if ab2.button("🗑️ Borrar", key=f"dal_{r['id']}"):
                                        try:
                                            supabase.table("inscripciones").delete().eq("id", r['id']).execute()
                                            st.success("Alumno eliminado del curso."); st.rerun()
                                        except Exception as e_err:
                                            st.error(f"Error: {e_err}")
                    except Exception as e:
                        st.error(f"Error al cargar alumnos: {e}")

    # =========================================================
    # TAB 4 — NOTAS
    # =========================================================
    with tabs[4]:
        st.subheader("📝 Notas y Calificaciones")
        if not mapa_cursos:
            no_encontrado("No hay cursos creados.")
        else:
            sub_nt = st.radio("Acción:", ["📋 Ver Notas por Curso", "✏️ Cargar Nota"], horizontal=True)
            if sub_nt == "📋 Ver Notas por Curso":
                col_f1, col_f2 = st.columns([2, 1])
                c_ver = col_f1.selectbox("Seleccione Curso:", ["---"] + list(mapa_cursos.keys()), key="nt_ver_sel")
                filtro_estado = col_f2.selectbox("Filtrar por estado:", ["Todos", "✅ Aprobados", "❌ Desaprobados"], key="filtro_estado_sel")
                busq_nota = st.text_input("🔍 Buscar alumno:", key="busq_nota_input")
                if c_ver != "---":
                    curso_data = mapa_cursos_data.get(c_ver, {})
                    nota_aprobacion = curso_data.get('nota_aprobacion')
                    if nota_aprobacion: st.caption(f"Nota de aprobación del curso: {nota_aprobacion}")
                    try:
                        res_al_v = supabase.table("inscripciones").select("id, alumnos(id, nombre, apellido, email)").eq("nombre_curso_materia", c_ver).not_.is_("alumno_id", "null").execute()
                        if not res_al_v.data:
                            no_encontrado("No hay alumnos inscriptos en este curso.")
                        else:
                            mostrados = 0
                            for r in res_al_v.data:
                                al_raw = r.get('alumnos')
                                al = al_raw[0] if isinstance(al_raw, list) and al_raw else al_raw
                                if al:
                                    if busq_nota.strip() and busq_nota.lower() not in al.get('nombre','').lower() and busq_nota.lower() not in al.get('apellido','').lower(): continue
                                    try:
                                        res_notas = supabase.table("notas").select("id, calificacion, created_at").eq("inscripcion_id", r['id']).order("created_at", desc=False).execute()
                                        notas = res_notas.data if res_notas.data else []
                                    except: notas = []
                                    filas_html = ""; valores = []
                                    for i, nt in enumerate(notas):
                                        val = float(nt['calificacion']); valores.append(val)
                                        fecha_fmt = datetime.datetime.fromisoformat(nt['created_at'][:10]).strftime('%d/%m/%Y')
                                        filas_html += f'<div class="nota-linea"><span>Nota {i+1} · {fecha_fmt}</span><span class="nota-badge {color_nota(val)}">{val}</span></div>'
                                    if not valores:
                                        if filtro_estado != "Todos": continue
                                        filas_html = '<div class="nota-linea"><span style="color:#555">Sin notas cargadas</span></div>'
                                        promedio_html = estado_html = ""
                                    else:
                                        promedio = round(sum(valores)/len(valores), 2)
                                        es_aprobado = nota_aprobacion is not None and promedio >= float(nota_aprobacion)
                                        if filtro_estado == "✅ Aprobados" and not es_aprobado: continue
                                        if filtro_estado == "❌ Desaprobados" and es_aprobado: continue
                                        promedio_html = f'<div class="promedio-linea"><span>Promedio</span><span class="nota-badge {color_nota(promedio)}">{promedio}</span></div>'
                                        estado_html = estado_aprobacion(promedio, nota_aprobacion)
                                    email_html = f'<div class="email-tag">✉️ {al.get("email","")}</div>' if al.get("email") else ""
                                    mostrados += 1
                                    st.markdown(f'<div class="alumno-block"><div class="nombre">👤 {al.get("apellido","").upper()}, {al.get("nombre","")} &nbsp; {estado_html if valores else ""}</div>{email_html}{filas_html}{promedio_html if valores else ""}</div>', unsafe_allow_html=True)
                                    if notas:
                                        for i, nt in enumerate(notas):
                                            col_n, col_d = st.columns([6, 1])
                                            col_n.caption(f"Nota {i+1}: {nt['calificacion']}")
                                            if col_d.button("🗑️", key=f"del_nota_{nt['id']}"):
                                                st.session_state[f'confirm_nota_{nt["id"]}'] = True; st.rerun()
                                            if st.session_state.get(f'confirm_nota_{nt["id"]}'):
                                                st.warning(f"⚠️ ¿Borrar Nota {i+1}: {nt['calificacion']}?")
                                                c1, c2 = st.columns(2)
                                                if c1.button("✅ Sí, borrar", key=f"conf_nota_{nt['id']}"):
                                                    supabase.table("notas").delete().eq("id", nt['id']).execute()
                                                    st.session_state[f'confirm_nota_{nt["id"]}'] = False
                                                    st.success("Nota eliminada."); st.rerun()
                                                if c2.button("❌ Cancelar", key=f"canc_nota_{nt['id']}"):
                                                    st.session_state[f'confirm_nota_{nt["id"]}'] = False; st.rerun()
                            if mostrados == 0:
                                no_encontrado("No hay alumnos que coincidan.")
                            else:
                                st.caption(f"Mostrando {mostrados} alumno/s")
                    except Exception as e:
                        st.error(f"Error: {e}")
            else:
                c_nt = st.selectbox("Seleccione Curso:", ["---"] + list(mapa_cursos.keys()), key="nt_carga_sel")
                busq_carga = st.text_input("🔍 Buscar alumno:", key="busq_carga_input")
                if st.session_state.get('ok_nota_agregada') is not None:
                    st.success(f"✅ Nota {st.session_state.ok_nota_agregada} agregada satisfactoriamente.")
                    st.session_state.ok_nota_agregada = None
                if c_nt != "---":
                    try:
                        res_al_n = supabase.table("inscripciones").select("id, alumnos(id, nombre, apellido, email)").eq("nombre_curso_materia", c_nt).not_.is_("alumno_id", "null").execute()
                        if not res_al_n.data:
                            no_encontrado("No hay alumnos inscriptos en este curso.")
                        else:
                            mostrados_carga = 0
                            for r in res_al_n.data:
                                al_raw = r.get('alumnos')
                                al = al_raw[0] if isinstance(al_raw, list) and al_raw else al_raw
                                if al:
                                    if busq_carga.strip() and busq_carga.lower() not in al.get('nombre','').lower() and busq_carga.lower() not in al.get('apellido','').lower(): continue
                                    mostrados_carga += 1
                                    try:
                                        res_notas = supabase.table("notas").select("id, calificacion, created_at").eq("inscripcion_id", r['id']).order("created_at", desc=False).execute()
                                        notas_existentes = res_notas.data if res_notas.data else []
                                    except: notas_existentes = []
                                    email_display = f'<br><span class="email-tag">✉️ {al.get("email","")}</span>' if al.get('email') else ''
                                    st.markdown(f'<div class="planilla-row">👤 {al.get("apellido","").upper()}, {al.get("nombre","")}{email_display}</div>', unsafe_allow_html=True)
                                    if notas_existentes:
                                        for i, nt in enumerate(notas_existentes):
                                            fecha_fmt = datetime.datetime.fromisoformat(nt['created_at'][:10]).strftime('%d/%m/%Y')
                                            col_n, col_d = st.columns([6, 1])
                                            col_n.markdown(f'<p class="nota-existente">Nota {i+1}: <b>{nt["calificacion"]}</b> · {fecha_fmt}</p>', unsafe_allow_html=True)
                                            if col_d.button("🗑️", key=f"del_nota_c_{nt['id']}"):
                                                st.session_state[f'confirm_nota_c_{nt["id"]}'] = True; st.rerun()
                                            if st.session_state.get(f'confirm_nota_c_{nt["id"]}'):
                                                st.warning(f"⚠️ ¿Borrar Nota {i+1}?")
                                                c1, c2 = st.columns(2)
                                                if c1.button("✅ Sí", key=f"conf_nota_c_{nt['id']}"):
                                                    supabase.table("notas").delete().eq("id", nt['id']).execute()
                                                    st.session_state[f'confirm_nota_c_{nt["id"]}'] = False
                                                    st.success("Nota eliminada."); st.rerun()
                                                if c2.button("❌ No", key=f"canc_nota_c_{nt['id']}"):
                                                    st.session_state[f'confirm_nota_c_{nt["id"]}'] = False; st.rerun()
                                    with st.form(f"nt_{r['id']}", clear_on_submit=True):
                                        nueva_nota = st.number_input("Nueva calificación:", 0.0, 10.0, value=0.0, step=0.1, key=f"ni_{r['id']}")
                                        if st.form_submit_button("💾 Agregar Nota"):
                                            try:
                                                supabase.table("notas").insert({"inscripcion_id": r['id'], "alumno_id": al['id'], "calificacion": nueva_nota}).execute()
                                                st.session_state.ok_nota_agregada = nueva_nota
                                                st.rerun()
                                            except Exception as e_err:
                                                st.error(f"Error: {e_err}")
                            if mostrados_carga == 0:
                                no_encontrado("No hay alumnos que coincidan.")
                    except Exception as e:
                        st.error(f"Error: {e}")

    # =========================================================
    # TAB 5 — ESTADÍSTICAS
    # =========================================================
    with tabs[5]:
        render_tab_estadisticas(u_data['id'], mapa_cursos, mapa_cursos_data)

    # =========================================================
    # TAB 6 — CONTADOR DE CLASES
    # =========================================================
    with tabs[6]:
        st.subheader("🔢 Contador de Clases")
        if not mapa_cursos:
            no_encontrado("No tenés cursos creados.")
        else:
            curso_cont = st.selectbox("Seleccione Curso:", ["---"] + list(mapa_cursos.keys()), key="cont_curso_sel")
            if curso_cont != "---":
                i_c = mapa_cursos[curso_cont]
                cd = mapa_cursos_data.get(curso_cont, {})
                hi_c = str(cd.get('hora_inicio', '') or '')[:5]
                hf_c = str(cd.get('hora_fin', '') or '')[:5]
                nota_ap_c = cd.get('nota_aprobacion')
                try:
                    res_count = supabase.table("inscripciones").select("id", count="exact").eq("nombre_curso_materia", curso_cont).not_.is_("alumno_id", "null").execute()
                    cant_alumnos = res_count.count if res_count.count else 0
                except: cant_alumnos = 0
                cant_clases, ultima_clase = get_stats_curso(i_c)
                ultima_html = f"Última clase: <b>{ultima_clase}</b>" if ultima_clase else "Sin clases registradas aún"
                st.markdown(f'''<div class="contador-card">
                    <div class="cc-nombre">📖 {curso_cont}</div>
                    <div class="cc-info">🕐 {format_horario(hi_c, hf_c)} &nbsp;·&nbsp; Alumnos: {cant_alumnos} &nbsp;·&nbsp; Aprobación: {nota_ap_c if nota_ap_c else "Sin definir"}</div>
                    <div class="cc-clases">📅 Clases dictadas: <b style="font-size:1.1rem;color:#4facfe">{cant_clases}</b> &nbsp;·&nbsp; {ultima_html}</div>
                </div>''', unsafe_allow_html=True)
            else:
                st.caption("📊 Resumen de todos los cursos:")
                for n_c, i_c in mapa_cursos.items():
                    cd = mapa_cursos_data.get(n_c, {})
                    hi_c = str(cd.get('hora_inicio', '') or '')[:5]
                    hf_c = str(cd.get('hora_fin', '') or '')[:5]
                    nota_ap_c = cd.get('nota_aprobacion')
                    try:
                        res_count = supabase.table("inscripciones").select("id", count="exact").eq("nombre_curso_materia", n_c).not_.is_("alumno_id", "null").execute()
                        cant_alumnos = res_count.count if res_count.count else 0
                    except: cant_alumnos = 0
                    cant_clases, ultima_clase = get_stats_curso(i_c)
                    ultima_html = f"Última: <b>{ultima_clase}</b>" if ultima_clase else "Sin clases aún"
                    st.markdown(f'''<div class="contador-card">
                        <div class="cc-nombre">📖 {n_c}</div>
                        <div class="cc-info">🕐 {format_horario(hi_c, hf_c)} &nbsp;·&nbsp; Alumnos: {cant_alumnos} &nbsp;·&nbsp; Aprobación: {nota_ap_c if nota_ap_c else "Sin definir"}</div>
                        <div class="cc-clases">📅 Clases dictadas: <b style="font-size:1.1rem;color:#4facfe">{cant_clases}</b> &nbsp;·&nbsp; {ultima_html}</div>
                    </div>''', unsafe_allow_html=True)

    # =========================================================
    # TAB 7 — CURSOS
    # =========================================================
    with tabs[7]:
        sub_cu = st.radio("Acción:", ["Mis Cursos", "Crear Nuevo Curso"], horizontal=True)
        if sub_cu == "Crear Nuevo Curso":
            if st.session_state.get('ok_curso_creado'):
                st.success(f"✅ Curso '{st.session_state.ok_curso_creado}' creado satisfactoriamente.")
                st.session_state.ok_curso_creado = None
            with st.form("new_c", clear_on_submit=True):
                mat = st.text_input("Nombre del Curso *")
                dias = st.multiselect("Días:", ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"])
                col_h1, col_h2 = st.columns(2)
                hora_ini = col_h1.text_input("Hora de inicio *", placeholder="hh:mm")
                hora_fin = col_h2.text_input("Hora de finalización *", placeholder="hh:mm")
                nota_ap = st.number_input("Nota de aprobación *", min_value=1.0, max_value=10.0, value=None, step=0.5, placeholder="Nota de aprobación")
                horas_catedra_new = None
                if es_universitario:
                    horas_catedra_new = st.number_input("Carga horaria semanal (hs cátedra) *", min_value=1, max_value=20, value=3, step=1)
                biblio = st.text_area("Bibliografía / Fotocopias (opcional)")
                url_campus = st.text_input("🌐 URL del Campus (opcional)", placeholder="https://...")
                nro_autom = st.text_input("🔢 Número de Automatriculación (opcional)")
                if st.form_submit_button("💾 CREAR CURSO"):
                    errores = []
                    if not mat.strip(): errores.append("El nombre del curso es obligatorio.")
                    if not dias: errores.append("Seleccioná al menos un día.")
                    if not hora_ini.strip(): errores.append("La hora de inicio es obligatoria.")
                    elif not validar_hora(hora_ini): errores.append("Hora de inicio inválida.")
                    if not hora_fin.strip(): errores.append("La hora de finalización es obligatoria.")
                    elif not validar_hora(hora_fin): errores.append("Hora de finalización inválida.")
                    if nota_ap is None: errores.append("La nota de aprobación es obligatoria.")
                    if errores:
                        for e in errores: st.error(e)
                    else:
                        info = f"{mat.strip()} ({', '.join(dias)}) | {hora_ini.strip()} → {hora_fin.strip()}"
                        try:
                            datos_curso = {
                                "profesor_id": u_data['id'], "nombre_curso_materia": info,
                                "anio_lectivo": 2026, "nota_aprobacion": nota_ap,
                                "bibliografia": biblio.strip() if biblio.strip() else None,
                                "hora_inicio": hora_ini.strip(), "hora_fin": hora_fin.strip(),
                                "url_campus": url_campus.strip() if url_campus.strip() else None,
                                "nro_automatriculacion": nro_autom.strip() if nro_autom.strip() else None,
                            }
                            if es_universitario and horas_catedra_new:
                                datos_curso["horas_catedra"] = horas_catedra_new
                            supabase.table("inscripciones").insert(datos_curso).execute()
                            st.session_state.ok_curso_creado = info
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error: {e}")
        else:
            if not mapa_cursos:
                no_encontrado("No tenés cursos creados todavía.")
            else:
                # Contador de cursos por año lectivo
                anio_lectivo_disp = 2026
                total_cursos_anio = len([c for c in mapa_cursos_data.values() if c.get('anio_lectivo') == anio_lectivo_disp])
                st.markdown(f'''<div class="contador-card" style="margin-bottom:16px;">
                    <div class="cc-nombre" style="font-family:\'Syne\',sans-serif;font-size:1rem;font-weight:800;letter-spacing:0.05em;">
                        Cantidad de Cursos de {anio_lectivo_disp}:
                        <span style="color:#4facfe;font-size:1.3rem;margin-left:8px;">{total_cursos_anio}</span>
                    </div>
                </div>''', unsafe_allow_html=True)
                if st.session_state.get('ok_curso_editado'):
                    st.success("✅ Curso actualizado satisfactoriamente.")
                    st.session_state.ok_curso_editado = False
                busq_curso = st.text_input("🔍 Buscar curso:", key="busq_curso")
                cursos_filtrados = {n: i for n, i in mapa_cursos.items() if not busq_curso.strip() or busq_curso.lower() in n.lower()}
                if busq_curso.strip() and not cursos_filtrados:
                    no_encontrado(f"No se encontró ningún curso con '{busq_curso}'.")
                else:
                    for n_c, i_c in cursos_filtrados.items():
                        curso_data = mapa_cursos_data.get(n_c, {})
                        nota_ap_cur = curso_data.get('nota_aprobacion')
                        biblio_cur = curso_data.get('bibliografia', '')
                        hi_cur = str(curso_data.get('hora_inicio', '') or '')[:5]
                        hf_cur = str(curso_data.get('hora_fin', '') or '')[:5]
                        horas_catedra_cur = curso_data.get('horas_catedra')
                        url_campus_cur = curso_data.get('url_campus', '') or ''
                        nro_autom_cur = curso_data.get('nro_automatriculacion', '') or ''
                        if st.session_state.editando_curso == i_c:
                            partes = n_c.split(" (")
                            mat_actual = partes[0] if partes else n_c
                            try: val_nota_edit = float(nota_ap_cur) if nota_ap_cur is not None else None
                            except: val_nota_edit = None
                            with st.form(f"edit_c_{i_c}"):
                                mat_e = st.text_input("Nombre del Curso:", value=mat_actual)
                                dias_e = st.multiselect("Días:", ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"])
                                col_h1e, col_h2e = st.columns(2)
                                hora_ini_e = col_h1e.text_input("Hora de inicio:", value=hi_cur, placeholder="hh:mm")
                                hora_fin_e = col_h2e.text_input("Hora de finalización:", value=hf_cur, placeholder="hh:mm")
                                nota_ap_e = st.number_input("Nota de aprobación *", min_value=1.0, max_value=10.0, value=val_nota_edit, step=0.5)
                                horas_cat_e = None
                                if es_universitario:
                                    horas_cat_e = st.number_input("Carga horaria semanal (hs cátedra):", min_value=1, max_value=20, value=int(horas_catedra_cur or 3), step=1)
                                biblio_e = st.text_area("Bibliografía:", value=biblio_cur or "")
                                url_campus_e = st.text_input("🌐 URL del Campus (opcional)", value=url_campus_cur, placeholder="https://...")
                                nro_autom_e = st.text_input("🔢 Número de Automatriculación (opcional)", value=nro_autom_cur)
                                col_c1, col_c2 = st.columns(2)
                                if col_c1.form_submit_button("💾 Guardar"):
                                    errores = []
                                    if hora_ini_e.strip() and not validar_hora(hora_ini_e): errores.append("Hora de inicio inválida.")
                                    if hora_fin_e.strip() and not validar_hora(hora_fin_e): errores.append("Hora de finalización inválida.")
                                    if errores:
                                        for e in errores: st.error(e)
                                    else:
                                        nuevo_nombre = f"{mat_e.strip()} ({', '.join(dias_e)}) | {hora_ini_e.strip()} → {hora_fin_e.strip()}" if dias_e else mat_e.strip()
                                        try:
                                            upd = {
                                                "nombre_curso_materia": nuevo_nombre, "nota_aprobacion": nota_ap_e,
                                                "bibliografia": biblio_e.strip() if biblio_e.strip() else None,
                                                "hora_inicio": hora_ini_e.strip() if hora_ini_e.strip() else None,
                                                "hora_fin": hora_fin_e.strip() if hora_fin_e.strip() else None,
                                                "url_campus": url_campus_e.strip() if url_campus_e.strip() else None,
                                                "nro_automatriculacion": nro_autom_e.strip() if nro_autom_e.strip() else None,
                                            }
                                            if es_universitario and horas_cat_e:
                                                upd["horas_catedra"] = horas_cat_e
                                            supabase.table("inscripciones").update(upd).eq("id", i_c).execute()
                                            st.session_state.editando_curso = None
                                            st.session_state.ok_curso_editado = True
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"Error: {e}")
                                if col_c2.form_submit_button("❌ Cancelar"):
                                    st.session_state.editando_curso = None; st.rerun()
                        else:
                            try:
                                res_count = supabase.table("inscripciones").select("id", count="exact").eq("nombre_curso_materia", n_c).not_.is_("alumno_id", "null").execute()
                                cant = res_count.count if res_count.count else 0
                            except: cant = 0
                            nota_display = nota_ap_cur if nota_ap_cur is not None else "Sin definir"
                            hs_display = f" &nbsp;·&nbsp; {horas_catedra_cur} hs/sem" if es_universitario and horas_catedra_cur else ""
                            biblio_html = f'<div class="biblio-box">📚 {biblio_cur}</div>' if biblio_cur else ""
                            st.markdown(
                                f'<div class="planilla-row">📖 {n_c}<br>'
                                f'<small style="color:#4facfe;">🕐 {format_horario(hi_cur, hf_cur)} &nbsp;·&nbsp; Alumnos: {cant} &nbsp;·&nbsp; Aprobación: {nota_display}{hs_display}</small>'
                                f'{biblio_html}</div>', unsafe_allow_html=True
                            )
                            cb1, cb2 = st.columns(2)
                            if cb1.button("✏️ Editar", key=f"ec_{i_c}"):
                                st.session_state.editando_curso = i_c; st.rerun()
                            if cb2.button("🗑️ Borrar", key=f"dc_{i_c}"):
                                try:
                                    supabase.table("inscripciones").delete().eq("id", i_c).execute()
                                    st.success("Curso eliminado."); st.rerun()
                                except Exception as e:
                                    st.error(f"Error: {e}")

    # =========================================================
    # TAB 8 — CALENDARIO ACADÉMICO
    # =========================================================
    with tabs[8]:
        st.subheader(f"📆 Calendario Académico {ANIO_ACTUAL}")
        render_seccion_calendario(u_data['sede'])

    # =========================================================
    # TAB 9 — IMPRESIÓN
    # =========================================================
    with tabs[9]:
        st.subheader("🖨️ Impresión y Exportación")
        if not mapa_cursos:
            no_encontrado("No tenés cursos creados.")
        else:
            col_i1, col_i2 = st.columns([2, 1])
            curso_imp = col_i1.selectbox("Seleccione Curso:", list(mapa_cursos.keys()), key="imp_curso")
            accion_imp = col_i2.selectbox("Acción:", ["📄 Exportar PDF", "📊 Exportar Excel", "🖨️ Imprimir"], key="imp_accion")
            if accion_imp == "📊 Exportar Excel":
                st.info("📊 El Excel incluirá el listado completo de alumnos con todas sus notas y promedios.")
                if st.button("⚙️ Generar Excel", type="primary", use_container_width=True):
                    if not EXCEL_OK:
                        st.error("La librería openpyxl no está instalada.")
                    else:
                        with st.spinner("Generando Excel..."):
                            inscripcion_id_imp = mapa_cursos[curso_imp]
                            curso_data_imp = mapa_cursos_data.get(curso_imp, {})
                            datos_imp = cargar_datos_por_nombre_curso(curso_imp, inscripcion_id_imp)
                            xlsx_bytes = generar_excel(u_data['sede'], curso_imp, curso_data_imp, datos_imp)
                            nombre_xlsx = f"ClassTrack_{u_data['sede']}_{curso_imp[:20].replace(' ','_')}_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
                            st.download_button(label="⬇️ Descargar Excel", data=xlsx_bytes, file_name=nombre_xlsx, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                            st.success("Excel generado.")
            else:
                st.markdown("**¿Qué incluir?**")
                col_c1, col_c2, col_c3, col_c4 = st.columns(4)
                inc_resumen = col_c1.checkbox("📊 Resumen del curso", value=True, key="inc_resumen")
                inc_alumnos = col_c2.checkbox("👥 Listado de alumnos", value=True, key="inc_alumnos")
                inc_notas = col_c3.checkbox("📝 Notas y promedios", value=True, key="inc_notas")
                inc_historial = col_c4.checkbox("📅 Historial de clases", value=False, key="inc_historial")
                if not any([inc_resumen, inc_alumnos, inc_notas, inc_historial]):
                    st.warning("Seleccioná al menos una sección para incluir.")
                else:
                    if st.button("⚙️ Generar documento", type="primary", use_container_width=True):
                        with st.spinner("Generando..."):
                            inscripcion_id_imp = mapa_cursos[curso_imp]
                            curso_data_imp = mapa_cursos_data.get(curso_imp, {})
                            datos_imp = cargar_datos_por_nombre_curso(curso_imp, inscripcion_id_imp)
                            if accion_imp == "📄 Exportar PDF":
                                pdf_bytes = generar_pdf(u_data['sede'], curso_imp, curso_data_imp, inc_alumnos, inc_notas, inc_historial, inc_resumen, datos_imp)
                                nombre_archivo = f"ClassTrack_{u_data['sede']}_{curso_imp[:20].replace(' ','_')}_{datetime.date.today().strftime('%Y%m%d')}.pdf"
                                st.download_button(label="⬇️ Descargar PDF", data=pdf_bytes, file_name=nombre_archivo, mime="application/pdf", use_container_width=True)
                                st.success("PDF generado.")
                            else:
                                html_imp = generar_html_impresion(u_data['sede'], curso_imp, curso_data_imp, inc_alumnos, inc_notas, inc_historial, inc_resumen, datos_imp)
                                html_encoded = html_imp.replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n").replace("\r", "")
                                js_code = f"<script>var w=window.open('','_blank');w.document.write('{html_encoded}');w.document.close();</script>"
                                components.html(js_code, height=0)
                                st.success("✅ Ventana de impresión abierta.")
            st.markdown("---")
            st.caption("💡 PDF ideal para enviar por mail · Excel para trabajar los datos · Imprimir abre el diálogo del navegador.")

    # =========================================================
    # TAB 10 — BACKUP
    # =========================================================
    with tabs[10]:
        st.subheader("🗄️ Backup y Restauración")

        st.markdown('''<div class="backup-aviso">
            ⚠️ <b>¿Por qué hacer backup?</b><br>
            Tus datos están en la nube, pero tener una copia local te protege ante cualquier imprevisto.
            Se recomienda hacer backup al menos <b>una vez por mes</b>.
        </div>''', unsafe_allow_html=True)

        # --- HISTORIAL DE BACKUPS ---
        historial_bk = get_historial_backups(u_data['id'])
        if historial_bk:
            st.markdown('<div class="backup-box"><div class="backup-titulo">📋 Historial de Backups</div>', unsafe_allow_html=True)
            for bk in historial_bk:
                try:
                    fecha_bk = datetime.datetime.fromisoformat(bk['fecha'].replace('Z', '+00:00'))
                    fecha_fmt = fecha_bk.strftime('%d/%m/%Y %H:%M')
                except: fecha_fmt = bk.get('fecha', '-')[:16]
                json_st = '<span class="backup-ok">✅ JSON</span>' if bk.get('json_descargado') else '<span class="backup-no">❌ JSON</span>'
                excel_st = '<span class="backup-ok">✅ Excel</span>' if bk.get('excel_descargado') else '<span class="backup-no">❌ Excel</span>'
                st.markdown(f'''<div class="backup-hist-row">
                    <span>📅 {fecha_fmt}</span>
                    <span style="color:#556">{bk.get("cant_cursos",0)} cursos · {bk.get("cant_alumnos",0)} alumnos · {bk.get("cant_clases",0)} clases</span>
                    <span>{json_st} &nbsp; {excel_st}</span>
                </div>''', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.caption("Aún no realizaste ningún backup.")

        st.markdown("---")

        # --- GENERAR BACKUP ---
        st.markdown('<div class="backup-box"><div class="backup-titulo">💾 Generar Backup Ahora</div>', unsafe_allow_html=True)

        if not st.session_state.backup_generado:
            if st.button("⚙️ Generar Backup Completo", type="primary", use_container_width=True):
                with st.spinner("Recopilando todos tus datos..."):
                    datos_bk = generar_datos_backup(u_data['id'], u_data['sede'])
                    json_bytes = datos_a_json_bytes(datos_bk)
                    excel_bytes = datos_a_excel_bytes(datos_bk, u_data['sede']) if EXCEL_OK else None
                    log_id = registrar_backup_log(u_data['id'], datos_bk)
                    st.session_state.backup_json_bytes = json_bytes
                    st.session_state.backup_excel_bytes = excel_bytes
                    st.session_state.backup_log_id = log_id
                    st.session_state.backup_generado = True
                    st.rerun()
        else:
            st.success("✅ Backup generado. Descargá los archivos:")
            fecha_nombre = datetime.datetime.now().strftime('%Y%m%d_%H%M')
            nombre_json = f"ClassTrack_{u_data['sede']}_{fecha_nombre}.json"
            nombre_excel = f"ClassTrack_{u_data['sede']}_{fecha_nombre}.xlsx"

            col_j, col_x = st.columns(2)
            with col_j:
                if st.download_button(
                    label="⬇️ Descargar JSON",
                    data=st.session_state.backup_json_bytes,
                    file_name=nombre_json,
                    mime="application/json",
                    use_container_width=True,
                    key="dl_json"
                ):
                    actualizar_backup_log(st.session_state.backup_log_id, json_desc=True)
                st.caption("📁 Guardalo en tu PC, pendrive o Google Drive. Es el archivo de respaldo técnico completo.")

            with col_x:
                if st.session_state.backup_excel_bytes:
                    if st.download_button(
                        label="⬇️ Descargar Excel",
                        data=st.session_state.backup_excel_bytes,
                        file_name=nombre_excel,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="dl_excel"
                    ):
                        actualizar_backup_log(st.session_state.backup_log_id, excel_desc=True)
                    st.caption("📊 Abrilo con Excel o Google Sheets para consultar tus datos fácilmente.")
                else:
                    st.warning("openpyxl no disponible.")

            if st.button("🔄 Generar nuevo backup", use_container_width=True):
                st.session_state.backup_generado = False
                st.session_state.backup_json_bytes = None
                st.session_state.backup_excel_bytes = None
                st.session_state.backup_log_id = None
                st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)

        # --- RESTAURAR DESDE BACKUP ---
        st.markdown('''<div class="restore-box">
            <div class="restore-titulo">⚠️ Restaurar desde Backup</div>''', unsafe_allow_html=True)
        st.markdown('''<div class="advertencia-box">
            ⚠️ <b>ADVERTENCIA:</b> La importación reemplaza <b>TODOS</b> tus datos actuales con los del backup.
            Esta acción <b>no puede deshacerse</b>.<br><br>
            Antes de importar, generá un backup de tu estado actual.
        </div>''', unsafe_allow_html=True)

        archivo_restore = st.file_uploader("Subir archivo JSON de backup:", type=['json'], key="restore_uploader")
        if archivo_restore:
            try:
                datos_restore = json.loads(archivo_restore.read().decode('utf-8'))
                meta = datos_restore.get('meta', {})
                cant_cursos_r = len(datos_restore.get('cursos', []))
                cant_alumnos_r = len(datos_restore.get('alumnos', []))
                cant_clases_r = len(datos_restore.get('historial_clases', []))
                fecha_bk_r = meta.get('fecha_backup', '-')[:16].replace('T', ' ')
                sede_bk_r = meta.get('sede', '-')
                st.markdown(f'''<div class="backup-box">
                    <div class="backup-titulo">📋 Preview del archivo</div>
                    <div class="backup-hist-row"><span>📅 Fecha backup</span><span>{fecha_bk_r}</span></div>
                    <div class="backup-hist-row"><span>🏫 Sede</span><span>{sede_bk_r.upper()}</span></div>
                    <div class="backup-hist-row"><span>📖 Cursos</span><span>{cant_cursos_r}</span></div>
                    <div class="backup-hist-row"><span>👥 Alumnos</span><span>{cant_alumnos_r}</span></div>
                    <div class="backup-hist-row"><span>📅 Clases</span><span>{cant_clases_r}</span></div>
                </div>''', unsafe_allow_html=True)

                if not st.session_state.get('confirmar_restore'):
                    if st.button("🔴 RESTAURAR ESTE BACKUP", type="primary", use_container_width=True):
                        st.session_state.confirmar_restore = True; st.rerun()
                else:
                    st.markdown('<div class="advertencia-box">⚠️ <b>ÚLTIMA CONFIRMACIÓN</b> — ¿Estás seguro? Se borrarán TODOS tus datos actuales.</div>', unsafe_allow_html=True)
                    col_sr, col_cr = st.columns(2)
                    if col_sr.button("✅ SÍ, RESTAURAR TODO", type="primary", use_container_width=True):
                        with st.spinner("Restaurando datos..."):
                            ok, msg = restaurar_desde_json(u_data['id'], datos_restore)
                            if ok:
                                registrar_backup_log(u_data['id'], datos_restore, json_desc=True)
                                st.session_state.confirmar_restore = False
                                st.success(f"✅ {msg}"); st.rerun()
                            else:
                                st.error(f"Error en la restauración: {msg}")
                    if col_cr.button("❌ CANCELAR", use_container_width=True):
                        st.session_state.confirmar_restore = False; st.rerun()
            except Exception as e:
                st.error(f"Error leyendo el archivo: {e}")

        st.markdown('</div>', unsafe_allow_html=True)

# ============================================================
# FIN PARTE 2 DE 2 — v297 completa ✅
# ============================================================
